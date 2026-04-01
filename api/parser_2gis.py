"""
ZEREK — Универсальный парсер 2ГИС
Парсит количество заведений и средние чеки по всем нишам и городам КЗ.
Запускать локально или на Railway (отдельный скрипт, не часть API).

Использование:
    python parser_2gis.py

Результат: data/parsed/2gis_data.xlsx
"""

import requests
import time
import json
import os
from datetime import datetime

# Попробуем импортировать openpyxl для сохранения в xlsx
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("openpyxl не установлен, результат будет в JSON")


# ═══════════════════════════════════════
# КОНФИГ
# ═══════════════════════════════════════

# Города КЗ с ID в 2ГИС (region_id)
CITIES = {
    'astana':       {'name': 'Астана',          '2gis_city': 'astana'},
    'almaty':       {'name': 'Алматы',          '2gis_city': 'almaty'},
    'shymkent':     {'name': 'Шымкент',         '2gis_city': 'shymkent'},
    'aktobe':       {'name': 'Актобе',          '2gis_city': 'aktobe'},
    'karaganda':    {'name': 'Караганда',       '2gis_city': 'karaganda'},
    'atyrau':       {'name': 'Атырау',          '2gis_city': 'atyrau'},
    'pavlodar':     {'name': 'Павлодар',        '2gis_city': 'pavlodar'},
    'aktau':        {'name': 'Актау',           '2gis_city': 'aktau'},
    'semey':        {'name': 'Семей',           '2gis_city': 'semey'},
    'kostanay':     {'name': 'Костанай',        '2gis_city': 'kostanay'},
    'uralsk':       {'name': 'Уральск',         '2gis_city': 'uralsk'},
    'taraz':        {'name': 'Тараз',           '2gis_city': 'taraz'},
    'ust_kamenogorsk': {'name': 'Усть-Каменогорск', '2gis_city': 'ust-kamenogorsk'},
    'petropavlovsk': {'name': 'Петропавловск',  '2gis_city': 'petropavlovsk'},
    'kyzylorda':    {'name': 'Кызылорда',       '2gis_city': 'kyzylorda'},
}

# Маппинг ниш на поисковые запросы 2ГИС
NICHE_QUERIES = {
    # Общепит
    'COFFEE':       ['Кофейни', 'Кофе с собой'],
    'BAKERY':       ['Пекарни', 'Пекарня'],
    'DONER':        ['Донерная', 'Шаурма', 'Шаверма'],
    'PIZZA':        ['Пиццерия', 'Доставка пиццы'],
    'SUSHI':        ['Суши-бар', 'Доставка суши'],
    'FASTFOOD':     ['Быстрое питание', 'Бургерная'],
    'CANTEEN':      ['Столовая', 'Бизнес-ланч'],
    
    # Красота
    'BARBER':       ['Барбершоп', 'Мужская стрижка'],
    'NAIL':         ['Маникюр', 'Ногтевой сервис'],
    'LASH':         ['Наращивание ресниц', 'Ресницы'],
    'SUGARING':     ['Шугаринг', 'Депиляция'],
    'BROW':         ['Брови', 'Оформление бровей'],
    'MASSAGE':      ['Массаж', 'Массажный салон'],
    
    # Здоровье
    'DENTAL':       ['Стоматология', 'Стоматологическая клиника'],
    'FITNESS':      ['Фитнес-клуб', 'Тренажёрный зал', 'Единоборства'],
    
    # Авто
    'CARWASH':      ['Автомойка', 'Мойка автомобилей'],
    'AUTOSERVICE':  ['Автосервис', 'СТО'],
    'TIRE':         ['Шиномонтаж'],
    
    # Торговля
    'GROCERY':      ['Продуктовый магазин', 'Минимаркет'],
    'PHARMA':       ['Аптека'],
    'FLOWERS':      ['Цветочный магазин', 'Цветы'],
    'FRUITSVEGS':   ['Овощи фрукты', 'Фруктовый магазин'],
    
    # Услуги
    'CLEAN':        ['Клининг', 'Уборка помещений'],
    'DRYCLEAN':     ['Химчистка', 'Прачечная'],
    'REPAIR_PHONE': ['Ремонт телефонов', 'Ремонт смартфонов'],
    'KINDERGARTEN': ['Частный детский сад', 'Детский сад'],
    'PVZ':          ['Пункт выдачи заказов', 'Ozon', 'Wildberries'],
    
    # Производство
    'SEMIFOOD':     ['Полуфабрикаты', 'Пельмени манты'],
    'CONFECTION':   ['Торты на заказ', 'Кондитерская'],
    'WATER':        ['Доставка воды', 'Питьевая вода'],
    
    # Другое
    'TAILOR':       ['Ателье', 'Пошив одежды'],
    'CYBERCLUB':    ['Компьютерный клуб', 'Кибер клуб'],
    'FURNITURE':    ['Мебель на заказ', 'Корпусная мебель'],
}


# ═══════════════════════════════════════
# ПАРСЕР
# ═══════════════════════════════════════

def parse_2gis_search(city_slug, query, delay=2.0):
    """
    Парсит страницу 2ГИС с поисковой выдачей.
    Возвращает: {count, avg_check, items: [{name, address, check, rating}]}
    
    Работает через HTML парсинг страницы 2gis.kz/{city}/search/{query}
    """
    url = f"https://2gis.kz/{city_slug}/search/{requests.utils.quote(query)}"
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'ru-RU,ru;q=0.9',
    }
    
    result = {
        'url': url,
        'count': 0,
        'avg_check': None,
        'checks': [],
        'items': [],
        'error': None,
    }
    
    try:
        resp = requests.get(url, headers=headers, timeout=15)
        text = resp.text
        
        # Парсим количество результатов
        # 2ГИС показывает "Места 109" в HTML
        import re
        
        # Ищем количество мест
        count_match = re.search(r'Места\s*(\d+)', text)
        if count_match:
            result['count'] = int(count_match.group(1))
        
        # Ищем средние чеки — формат "Чек XXXX тнг"
        checks = re.findall(r'Чек\s+(\d+)\s*тнг', text)
        if checks:
            checks_int = [int(c) for c in checks]
            result['checks'] = checks_int
            result['avg_check'] = int(sum(checks_int) / len(checks_int))
        
        # Ищем названия заведений и рейтинги
        # Это упрощённый парсер — 2ГИС рендерит через JS, 
        # поэтому полные данные через requests не получить.
        # Для полного парсинга нужен Playwright/Selenium.
        
    except requests.exceptions.Timeout:
        result['error'] = 'timeout'
    except Exception as e:
        result['error'] = str(e)
    
    time.sleep(delay)  # Задержка между запросами
    return result


def run_parser(cities=None, niches=None, delay=2.0):
    """
    Запускает парсинг по всем городам и нишам.
    cities: список city_id или None (все)
    niches: список niche_id или None (все)
    """
    if cities is None:
        cities = list(CITIES.keys())
    if niches is None:
        niches = list(NICHE_QUERIES.keys())
    
    total = len(cities) * len(niches)
    results = {}
    done = 0
    
    print(f"Парсинг 2ГИС: {len(cities)} городов × {len(niches)} ниш = {total} запросов")
    print(f"Задержка: {delay} сек. Примерное время: {total * delay / 60:.0f} мин")
    print("=" * 60)
    
    for city_id in cities:
        city = CITIES[city_id]
        city_slug = city['2gis_city']
        results[city_id] = {}
        
        for niche_id in niches:
            queries = NICHE_QUERIES[niche_id]
            
            # Парсим первый запрос (основной)
            main_query = queries[0]
            r = parse_2gis_search(city_slug, main_query, delay=delay)
            
            results[city_id][niche_id] = {
                'city': city['name'],
                'niche': niche_id,
                'query': main_query,
                'count': r['count'],
                'avg_check': r['avg_check'],
                'checks_found': len(r['checks']),
                'checks': r['checks'][:10],  # первые 10 чеков
                'error': r['error'],
            }
            
            done += 1
            status = '✓' if not r['error'] else '✗'
            check_str = f"чек ~{r['avg_check']} тг" if r['avg_check'] else "чек —"
            print(f"  [{done}/{total}] {status} {city['name']:15s} {niche_id:15s} → {r['count']:4d} мест, {check_str}")
    
    return results


def save_results(results, output_dir='data/parsed'):
    """Сохраняет результаты в xlsx и json."""
    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    
    # JSON
    json_path = os.path.join(output_dir, f'2gis_{ts}.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"\nJSON: {json_path}")
    
    # XLSX
    if HAS_OPENPYXL:
        xlsx_path = os.path.join(output_dir, f'2gis_{ts}.xlsx')
        wb = Workbook()
        
        # Лист 1: Сводная по количеству заведений
        ws1 = wb.active
        ws1.title = "Количество"
        niches = list(NICHE_QUERIES.keys())
        cities = list(results.keys())
        
        # Заголовки
        ws1.cell(row=1, column=1, value='Город / Ниша')
        for j, niche in enumerate(niches, 2):
            ws1.cell(row=1, column=j, value=niche)
        
        for i, city_id in enumerate(cities, 2):
            ws1.cell(row=i, column=1, value=CITIES[city_id]['name'])
            for j, niche in enumerate(niches, 2):
                data = results.get(city_id, {}).get(niche, {})
                ws1.cell(row=i, column=j, value=data.get('count', 0))
        
        # Лист 2: Средние чеки
        ws2 = wb.create_sheet("Средние чеки")
        ws2.cell(row=1, column=1, value='Город / Ниша')
        for j, niche in enumerate(niches, 2):
            ws2.cell(row=1, column=j, value=niche)
        
        for i, city_id in enumerate(cities, 2):
            ws2.cell(row=i, column=1, value=CITIES[city_id]['name'])
            for j, niche in enumerate(niches, 2):
                data = results.get(city_id, {}).get(niche, {})
                ws2.cell(row=i, column=j, value=data.get('avg_check'))
        
        # Лист 3: Детали
        ws3 = wb.create_sheet("Детали")
        ws3.append(['Город', 'Ниша', 'Запрос', 'Количество', 'Ср. чек', 'Чеки найдены', 'Ошибка'])
        for city_id, city_data in results.items():
            for niche_id, data in city_data.items():
                ws3.append([
                    data.get('city', ''),
                    data.get('niche', ''),
                    data.get('query', ''),
                    data.get('count', 0),
                    data.get('avg_check'),
                    data.get('checks_found', 0),
                    data.get('error', ''),
                ])
        
        wb.save(xlsx_path)
        print(f"XLSX: {xlsx_path}")
    
    return json_path


# ═══════════════════════════════════════
# ЗАПУСК
# ═══════════════════════════════════════

if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='ZEREK 2GIS Parser')
    parser.add_argument('--cities', nargs='+', default=None, help='Список городов (city_id)')
    parser.add_argument('--niches', nargs='+', default=None, help='Список ниш (niche_id)')
    parser.add_argument('--delay', type=float, default=2.0, help='Задержка между запросами (сек)')
    parser.add_argument('--output', default='data/parsed', help='Папка для результатов')
    parser.add_argument('--quick', action='store_true', help='Быстрый тест: 3 города × 3 ниши')
    
    args = parser.parse_args()
    
    if args.quick:
        args.cities = ['astana', 'almaty', 'aktobe']
        args.niches = ['COFFEE', 'BAKERY', 'CARWASH']
    
    print("ZEREK 2GIS Parser")
    print(f"Города: {args.cities or 'все'}")
    print(f"Ниши: {args.niches or 'все'}")
    print()
    
    results = run_parser(
        cities=args.cities,
        niches=args.niches,
        delay=args.delay,
    )
    
    save_results(results, args.output)
    print("\nГотово!")
