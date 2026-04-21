"""
ZEREK — Генератор финансовой модели из шаблона.
Берёт шаблон Адиля, подставляет данные из анкеты/движка.
v2: исправлен PARAM_MAP под реальную структуру шаблона,
    safe_write для merged cells.
v3: дефолты параметров и сезонность вынесены в
    config/finmodel_defaults.yaml. Сама логика generate_finmodel
    не переписывалась — только источник дефолтных значений.
"""
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from datetime import datetime
import os


def _load_finmodel_defaults() -> dict:
    """Читает config/finmodel_defaults.yaml. Возвращает {} при ошибке."""
    path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "config", "finmodel_defaults.yaml",
    )
    try:
        import yaml
        with open(path, "r", encoding="utf-8") as fh:
            return (yaml.safe_load(fh) or {}).get("finmodel", {}) or {}
    except Exception:
        return {}


_FM_DEFAULTS = _load_finmodel_defaults()
_FM_OPEX_DEFAULTS = _FM_DEFAULTS.get("opex", {}) or {}
_FM_FOT_DEFAULTS = _FM_DEFAULTS.get("fot", {}) or {}
_FM_CAPEX_DEFAULTS = _FM_DEFAULTS.get("capex", {}) or {}
_FM_CREDIT_DEFAULTS = _FM_DEFAULTS.get("credit", {}) or {}
_FM_GROWTH_DEFAULTS = _FM_DEFAULTS.get("growth", {}) or {}
_FM_SEASONALITY_DEFAULT = _FM_DEFAULTS.get(
    "default_seasonality",
    [0.85, 0.85, 0.90, 1.00, 1.05, 1.10, 1.10, 1.05, 1.00, 0.95, 0.95, 1.20],
)


# Карта ячеек ПАРАМЕТРЫ — СТРОГО по реальному шаблону Адиля
# row → (column, param_key)
PARAM_MAP = {
    # 0. ОБЩИЕ (строки 5-9)
    5:  'entity_type',     # ИП / ТОО
    6:  'tax_regime',      # УСН / ОУР
    7:  'nds_payer',       # Нет / Да
    8:  'tax_rate',        # 0.03
    9:  'horizon',         # 36
    # 1. ВЫРУЧКА (строки 12-16)
    12: 'check_med',       # Средний чек
    13: 'traffic_med',     # Клиентов/день
    14: 'work_days',       # Рабочих дней
    15: 'traffic_growth',  # Рост трафика %
    16: 'check_growth',    # Рост чека %
    # 2. ПРЯМЫЕ РАСХОДЫ (строки 19-20)
    19: 'cogs_pct',        # Себестоимость %
    20: 'loss_pct',        # Потери %
    # 3. ПОСТОЯННЫЕ РАСХОДЫ (строки 23-34)
    23: 'rent',            # Аренда
    # 24 — площадь удалена из шаблона
    25: 'fot_gross',       # ФОТ Gross
    26: 'headcount',       # Кол-во сотрудников
    27: 'utilities',       # Коммунальные
    28: 'marketing',       # Маркетинг
    29: 'consumables',     # Расходники
    30: 'software',        # Софт/касса
    31: 'other',           # Прочие
    # 32-34 — ссылки на ДОПУЩЕНИЯ, не трогаем
    # 4. СТАРТОВЫЕ ИНВЕСТИЦИИ (строки 37-40)
    37: 'capex',           # CAPEX
    38: 'deposit_months',  # Депозит мес
    39: 'working_cap',     # Оборотный капитал
    40: 'amort_years',     # Амортизация лет
    # 5. КРЕДИТ (строки 46-48)
    46: 'credit_amount',   # Сумма кредита
    47: 'credit_rate',     # Ставка
    48: 'credit_term',     # Срок мес
    # 6. ДИСКОНТИРОВАНИЕ (строка 52)
    52: 'wacc',            # WACC
}


def safe_write(ws, row, col, value):
    """Безопасная запись в ячейку — пропускает MergedCell."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return False
    cell.value = value
    return True


def safe_write_addr(ws, addr, value):
    """Безопасная запись по адресу (A1, B5 и т.д.)."""
    cell = ws[addr]
    if isinstance(cell, MergedCell):
        return False
    cell.value = value
    return True


def generate_finmodel(
    template_path: str,
    params: dict,
    seasonality: list = None,
    output_path: str = None,
    eq_note: str = '',
) -> str:
    if seasonality is None:
        seasonality = _FM_SEASONALITY_DEFAULT

    # Дефолты — из config/finmodel_defaults.yaml (с жёсткими fallback-значениями
    # на случай отсутствия yaml-файла).
    defaults = {
        'entity_type':    _FM_DEFAULTS.get('entity_type', 'ИП'),
        'tax_regime':     _FM_DEFAULTS.get('tax_regime', 'УСН'),
        'nds_payer':      _FM_DEFAULTS.get('nds_payer', 'Нет'),
        'tax_rate':       _FM_DEFAULTS.get('tax_rate', 0.03),
        'horizon':        _FM_DEFAULTS.get('horizon_months', 36),
        'check_med':      _FM_DEFAULTS.get('check_med', 1400),
        'traffic_med':    _FM_DEFAULTS.get('traffic_med', 70),
        'work_days':      _FM_DEFAULTS.get('work_days', 30),
        'traffic_growth': _FM_GROWTH_DEFAULTS.get('traffic_yr', 0.07),
        'check_growth':   _FM_GROWTH_DEFAULTS.get('check_yr', 0.08),
        'cogs_pct':       _FM_DEFAULTS.get('cogs_pct', 0.35),
        'loss_pct':       _FM_DEFAULTS.get('loss_pct', 0.03),
        'rent':           _FM_OPEX_DEFAULTS.get('rent', 70000),
        'fot_gross':      _FM_FOT_DEFAULTS.get('fot_gross', 200000),
        'headcount':      _FM_FOT_DEFAULTS.get('headcount', 2),
        'utilities':      _FM_OPEX_DEFAULTS.get('utilities', 15000),
        'marketing':      _FM_OPEX_DEFAULTS.get('marketing', 50000),
        'consumables':    _FM_OPEX_DEFAULTS.get('consumables', 3500),
        'software':       _FM_OPEX_DEFAULTS.get('software', 5000),
        'other':          _FM_OPEX_DEFAULTS.get('other', 10000),
        'capex':          _FM_CAPEX_DEFAULTS.get('default_kzt', 1500000),
        'deposit_months': _FM_CAPEX_DEFAULTS.get('deposit_months', 2),
        'working_cap':    _FM_CAPEX_DEFAULTS.get('working_cap_kzt', 1000000),
        'amort_years':    _FM_CAPEX_DEFAULTS.get('amort_years', 7),
        'credit_amount':  _FM_CREDIT_DEFAULTS.get('default_amount', 0),
        'credit_rate':    _FM_CREDIT_DEFAULTS.get('rate', 0.22),
        'credit_term':    _FM_CREDIT_DEFAULTS.get('term_months', 36),
        'wacc':           _FM_DEFAULTS.get('wacc', 0.20),
    }

    for k, v in defaults.items():
        if k not in params or params[k] is None:
            params[k] = v

    wb = load_workbook(template_path)
    P = "'⚙️ ПАРАМЕТРЫ'"
    today = datetime.now().strftime('%d.%m.%Y')

    # ── 1. Подставляем ПАРАМЕТРЫ ──
    ws = wb['⚙️ ПАРАМЕТРЫ']
    for row, key in PARAM_MAP.items():
        value = params.get(key)
        if value is not None:
            safe_write(ws, row, 3, value)

    # Дисклеймер CAPEX
    if eq_note:
        safe_write(ws, 37, 5, f'💡 {eq_note}')

    # Business name + city в НАВИГАЦИЮ строка 1-2
    biz_name = params.get('business_name', 'Бизнес')
    city_name = params.get('city', '')
    
    # Дата в футере
    safe_write(ws, 55, 2, f'ZEREK Financial Model | {today} | @zerekai_bot')

    # ── 2. Подставляем ДОПУЩЕНИЯ (сезонность) ──
    ws2 = wb['📊 ДОПУЩЕНИЯ']
    for j, coef in enumerate(seasonality[:12]):
        safe_write(ws2, 5, 3 + j, coef)

    # ── 3. Динамические заголовки ──
    # Используем прямые строки вместо формул — надёжнее для merged cells
    title_main = f'ФИНАНСОВАЯ МОДЕЛЬ — {biz_name.upper()}'
    title_sub = f'Город: {city_name}  |  Горизонт: {params.get("horizon", 36)} мес  |  ZEREK  |  {today}'
    
    sheets_titles = {
        '📋 НАВИГАЦИЯ': {'A1': title_main, 'A2': title_sub},
        '⚙️ ПАРАМЕТРЫ': {'A1': f'ПАРАМЕТРЫ — {biz_name.upper()}'},
        '📊 ДОПУЩЕНИЯ': {'A1': f'ДОПУЩЕНИЯ — {biz_name.upper()}'},
        '🎯 ДАШБОРД': {'A1': f'ДАШБОРД — {biz_name.upper()} ({city_name})'},
        '📈 P&L': {'A1': f'P&L — {biz_name.upper()} ({city_name})'},
        '💰 CASH FLOW': {'A1': f'CASH FLOW — {biz_name.upper()} ({city_name})'},
        '📑 БАЛАНС': {'A1': f'БАЛАНС — {biz_name.upper()}'},
    }

    for sheet_name, cells in sheets_titles.items():
        if sheet_name in wb.sheetnames:
            ws_t = wb[sheet_name]
            for addr, formula in cells.items():
                # Для merged cells — нужно найти "главную" ячейку merge
                cell = ws_t[addr]
                if isinstance(cell, MergedCell):
                    # Найдём merge range, содержащий эту ячейку
                    for mr in ws_t.merged_cells.ranges:
                        if addr in mr:
                            # Пишем в верхнюю левую ячейку merge
                            top_left = mr.min_row, mr.min_col
                            ws_t.cell(row=top_left[0], column=top_left[1]).value = formula
                            break
                else:
                    cell.value = formula

    # ── 4. Сохраняем ──
    if output_path is None:
        biz = params.get('entity_type', 'Бизнес')
        city = params.get('city', 'Город')
        output_path = f'ZEREK_FinModel_{biz}_{city}.xlsx'

    wb.save(output_path)
    return output_path


def generate_from_quickcheck(template_path: str, qc_result: dict, output_path: str = None) -> str:
    """Генерирует финмодель из результата Quick Check v3."""
    inp = qc_result.get('input', {})
    fin = qc_result.get('financials', {})
    sf = qc_result.get('staff', {})
    cap = qc_result.get('capex', {})
    tx = qc_result.get('tax', {})

    # CAPEX
    bk = cap.get('breakdown', {})
    capex_total = sum(bk.get(k, 0) or 0 for k in
        ['equipment', 'renovation', 'furniture', 'first_stock', 'permits_sez'])

    params = {
        'entity_type': 'ИП',
        'tax_regime': tx.get('regime', 'УСН'),
        'nds_payer': 'Нет',
        'tax_rate': (tx.get('rate_pct', 3) or 3) / 100,
        'check_med': fin.get('check_med', 1400),
        'traffic_med': fin.get('traffic_med', 70),
        'work_days': 30,
        'cogs_pct': fin.get('cogs_pct', 0.35),
        'loss_pct': fin.get('loss_pct', 0.03),
        'rent': fin.get('rent_month', 70000),
        'fot_gross': sf.get('fot_gross_med', sf.get('fot_net_med', 200000)),
        'headcount': sf.get('headcount', 2),
        'utilities': fin.get('utilities', 15000),
        'marketing': fin.get('marketing', 50000),
        'consumables': fin.get('consumables', 3500),
        'software': fin.get('software', 5000),
        'other': fin.get('transport', 10000),
        'capex': capex_total or cap.get('capex_med', 1500000),
        'deposit_months': fin.get('deposit_months', 2),
        'working_cap': bk.get('working_cap', 1000000) or 1000000,
    }

    eq_note = cap.get('eq_note', '')

    return generate_finmodel(template_path, params, output_path=output_path, eq_note=eq_note)
