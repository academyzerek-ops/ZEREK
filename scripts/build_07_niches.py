"""
Builds data/kz/07_niches.xlsx — adaptive-survey config for all target niches.

Sheets:
  1. «Ниши»              — one row per niche, 10 config columns, header on row 6 (header=5)
  2. «Специфичные вопросы» — catalog of question_id → text/options/applies_to_niches

Run:  python3 scripts/build_07_niches.py
"""
from __future__ import annotations

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ---------------------------------------------------------------------------
# Target roster (authoritative — prompt-supplied). Used as the canonical set
# of niche_ids; superset of what exists as format xlsx / insight files today.
# ---------------------------------------------------------------------------
TARGET_NICHES = [
    "BAKERY", "BARBER", "BEAUTY", "BROW", "BUBBLETEA", "BUILDMAT",
    "CANTEEN", "CARGO", "CARPETCLEAN", "CARWASH", "CATERING", "CLEAN",
    "COFFEE", "COMPCLUB", "CONFECTION", "COSMETOLOGY",
    "DENTAL", "DETAILING", "DONER", "DRIVING", "DRYCLEAN",
    "FASTFOOD", "FITNESS", "FLOWERS", "FRUITSVEGS", "FURNITURE",
    "GROCERY", "HOTEL",
    "KIDSCENTER", "KINDERGARTEN",
    "LANGUAGES", "LASH", "LAUNDRY", "LOFTFURNITURE",
    "MANICURE", "MASSAGE", "MEATSHOP",
    "NOTARY", "EVALUATION",
    "OPTICS",
    "PETSHOP", "PHARMACY", "PHOTO", "PIZZA", "PRINTING", "PVZ",
    "REALTOR", "REPAIR_PHONE",
    "SEMIFOOD", "SUGARING", "SUSHI",
    "TAILOR", "TIRESERVICE",
    "WATERPLANT",
    "YOGA",
    "ACCOUNTING", "AUTOPARTS", "AUTOSERVICE",
]

# Human-readable names (RU)
NICHE_NAMES = {
    "BAKERY": "Пекарня", "BARBER": "Барбершоп", "BEAUTY": "Салон красоты",
    "BROW": "Брови и ресницы", "BUBBLETEA": "Островок напитков",
    "BUILDMAT": "Строительные материалы", "CANTEEN": "Столовая",
    "CARGO": "Грузоперевозки", "CARPETCLEAN": "Чистка ковров",
    "CARWASH": "Автомойка", "CATERING": "Кейтеринг", "CLEAN": "Клининг",
    "COFFEE": "Кофейня", "COMPCLUB": "Компьютерный клуб",
    "CONFECTION": "Кондитерская", "COSMETOLOGY": "Косметология",
    "DENTAL": "Стоматология", "DETAILING": "Детейлинг", "DONER": "Донерная",
    "DRIVING": "Автошкола", "DRYCLEAN": "Химчистка одежды",
    "FASTFOOD": "Фастфуд", "FITNESS": "Фитнес-клуб",
    "FLOWERS": "Цветочный магазин", "FRUITSVEGS": "Овощи и фрукты",
    "FURNITURE": "Производство корпусной мебели",
    "GROCERY": "Продуктовый магазин", "HOTEL": "Отель / гостиница",
    "KIDSCENTER": "Детский развивающий центр", "KINDERGARTEN": "Частный детский сад",
    "LANGUAGES": "Языковая школа", "LASH": "Ресницы",
    "LAUNDRY": "Прачечная", "LOFTFURNITURE": "Производство лофт-мебели",
    "MANICURE": "Маникюр", "MASSAGE": "Массажный салон", "MEATSHOP": "Мясная лавка",
    "NOTARY": "Нотариус", "EVALUATION": "Оценочные услуги",
    "OPTICS": "Оптика", "PETSHOP": "Зоомагазин",
    "PHARMACY": "Аптека", "PHOTO": "Фотостудия", "PIZZA": "Пиццерия",
    "PRINTING": "Типография / полиграфия", "PVZ": "Пункт выдачи заказов",
    "REALTOR": "Риэлторские услуги", "REPAIR_PHONE": "Ремонт телефонов",
    "SEMIFOOD": "Полуфабрикаты", "SUGARING": "Шугаринг", "SUSHI": "Суши",
    "TAILOR": "Ателье", "TIRESERVICE": "Шиномонтаж",
    "WATERPLANT": "Розлив воды", "YOGA": "Йога-студия",
    "ACCOUNTING": "Бухгалтерский аутсорсинг", "AUTOPARTS": "Автозапчасти",
    "AUTOSERVICE": "Автосервис",
}

# ---------------------------------------------------------------------------
# Rule sets from the prompt (Step 3)
# ---------------------------------------------------------------------------
LICENSE_MANDATORY = {
    "NOTARY": "Лицензия Минюста РК на нотариальную деятельность",
    "EVALUATION": "Лицензия уполномоченного органа на оценочную деятельность",
    "DENTAL": "Медицинская лицензия Минздрава РК",
    "COSMETOLOGY": "Медицинская лицензия (для инвазивной косметологии)",
    "KINDERGARTEN": "Образовательная лицензия Управления образования",
    "DRIVING": "Лицензия на обучение вождению",
    "PHARMACY": "Фармацевтическая лицензия Минздрава РК",
}
LICENSE_OPTIONAL = {
    # KIDSCENTER / LANGUAGES лицензия нужна ТОЛЬКО если работают с документом об образовании
    "KIDSCENTER": "Образовательная лицензия — если досадиковая группа или подготовка к школе",
    "LANGUAGES": "Образовательная лицензия — если планируется формальное обучение с аттестатами",
}

# Где уместна градация эконом/стандарт/премиум
CLASS_GRADES_YES = {
    "BARBER", "BEAUTY", "MANICURE", "BROW", "LASH", "SUGARING", "MASSAGE", "COSMETOLOGY",
    "COFFEE", "PIZZA", "SUSHI", "BUBBLETEA",
    "HOTEL", "DETAILING", "CARWASH", "FITNESS", "YOGA",
    "FURNITURE", "LOFTFURNITURE", "DENTAL", "OPTICS",
    "CATERING", "FASTFOOD", "CANTEEN", "SEMIFOOD", "DONER", "BAKERY",
    "FLOWERS", "PHOTO", "TAILOR",
}

# Где владелец может сам работать специалистом (yes)
SELF_OP_YES = {
    # Beauty
    "BARBER", "MANICURE", "BROW", "LASH", "SUGARING", "MASSAGE", "COSMETOLOGY", "BEAUTY",
    # One-person service / craft
    "PHOTO", "TAILOR", "ACCOUNTING", "REALTOR", "EVALUATION", "NOTARY", "LANGUAGES",
    # Small service crafts
    "PVZ", "REPAIR_PHONE", "DETAILING", "CARGO", "DRIVING",
    "CARPETCLEAN", "CLEAN", "DRYCLEAN", "LAUNDRY", "PRINTING",
}

# Локации по нишам (allowed_location_types)
LOCATION_MAP = {
    # Только возле дороги / улица
    "AUTOSERVICE":   ("highway,street,own_building",                    "street"),
    "TIRESERVICE":   ("highway,street,own_building",                    "street"),
    "CARWASH":       ("highway,street,own_building",                    "street"),
    "DETAILING":     ("highway,street,own_building",                    "street"),
    # Масс-рынок сервис, где важна видимость
    "BARBER":        ("tc,street,residential_complex",                  "street"),
    "MANICURE":      ("tc,street,residential_complex,home",             "street"),
    "BROW":          ("tc,street,residential_complex,home",             "street"),
    "LASH":          ("tc,street,residential_complex,home",             "street"),
    "SUGARING":      ("tc,street,residential_complex,home",             "street"),
    "BEAUTY":        ("tc,street,residential_complex",                  "street"),
    "COSMETOLOGY":   ("street,residential_complex,business_center",     "street"),
    "MASSAGE":       ("tc,street,residential_complex,home",             "street"),
    "OPTICS":        ("tc,street,residential_complex",                  "tc"),
    # Трафик в ТЦ / retail food
    "REPAIR_PHONE":  ("tc,street,market",                               "tc"),
    "PVZ":           ("tc,street,residential_complex,market",           "street"),
    "FLOWERS":       ("tc,street,residential_complex,market",           "street"),
    "BUBBLETEA":     ("tc,street,market",                               "tc"),
    "FASTFOOD":      ("tc,street,market,residential_complex",           "street"),
    "DONER":         ("street,tc,market,residential_complex",           "street"),
    "COFFEE":        ("tc,street,residential_complex,business_center",  "street"),
    "PIZZA":         ("street,tc,residential_complex",                  "street"),
    "SUSHI":         ("street,tc,residential_complex",                  "street"),
    "CANTEEN":       ("street,business_center,own_building",            "business_center"),
    "BAKERY":        ("street,tc,residential_complex",                  "street"),
    "PETSHOP":       ("street,tc,residential_complex,market",           "street"),
    # Производство / склад
    "WATERPLANT":    ("street,own_building",                            "own_building"),
    "CONFECTION":    ("street,own_building,residential_complex",        "street"),
    "CATERING":      ("street,own_building",                            "own_building"),
    "FURNITURE":     ("street,own_building",                            "own_building"),
    "LOFTFURNITURE": ("street,own_building",                            "own_building"),
    "SEMIFOOD":      ("street,own_building",                            "street"),
    "MEATSHOP":      ("street,market,residential_complex",              "street"),
    "PRINTING":      ("street,business_center,own_building",            "street"),
    # Спальный район для «своих»
    "KIDSCENTER":    ("street,residential_area,residential_complex",    "residential_area"),
    "KINDERGARTEN":  ("street,residential_area,residential_complex,own_building", "residential_area"),
    "GROCERY":       ("street,residential_area,residential_complex,market", "residential_area"),
    "FRUITSVEGS":    ("street,residential_area,market,residential_complex", "street"),
    "PHARMACY":      ("street,residential_area,residential_complex,tc",   "street"),
    "LANGUAGES":     ("street,residential_area,residential_complex,home,business_center", "residential_area"),
    "YOGA":          ("street,residential_area,residential_complex,business_center", "residential_area"),
    "FITNESS":       ("street,residential_area,own_building,tc",        "residential_area"),
    "COMPCLUB":      ("street,residential_area,residential_complex",    "residential_area"),
    # Отдельные здания
    "HOTEL":         ("own_building,street",                            "own_building"),
    "DENTAL":        ("own_building,street,business_center,residential_complex", "street"),
    # Офис (B2B)
    "ACCOUNTING":    ("business_center,street,home",                    "business_center"),
    "REALTOR":       ("business_center,street,home",                    "business_center"),
    "NOTARY":        ("business_center,street",                         "business_center"),
    "EVALUATION":    ("business_center,street,home",                    "business_center"),
    # Дома / из дома
    "CLEAN":         ("home,street,business_center",                    "home"),
    "CARGO":         ("home,own_building,street",                       "home"),
    "PHOTO":         ("home,street,residential_complex",                "street"),
    "TAILOR":        ("home,street,residential_complex,tc",             "street"),
    "DRYCLEAN":      ("street,residential_complex,tc",                  "street"),
    "CARPETCLEAN":   ("own_building,street",                            "own_building"),
    "LAUNDRY":       ("street,residential_complex,business_center",     "street"),
    # Розница прочая
    "AUTOPARTS":     ("street,market,own_building",                     "street"),
    "BUILDMAT":      ("street,own_building,market",                     "own_building"),
    "DRIVING":       ("street,own_building,business_center",            "street"),
}

# «Где можно из дома» — если пользователь выбрал home, площадь скрывается
AREA_HIDDEN_IF_HOME = {
    "LANGUAGES", "PHOTO", "ACCOUNTING", "TAILOR", "REALTOR", "CLEAN", "CARGO",
    "MANICURE", "BROW", "LASH", "SUGARING", "MASSAGE",  # beauty-mode-на-дому
    "EVALUATION",
}

# Персонал: обязательно штат с первого дня
STAFF_HIDDEN_MANDATORY_TEAM = {
    "KINDERGARTEN", "KIDSCENTER",
    "DENTAL", "PHARMACY",
    "HOTEL", "FITNESS", "COMPCLUB",
    "WATERPLANT", "CONFECTION", "BAKERY", "MEATSHOP",
    "AUTOSERVICE", "TIRESERVICE",
    "CATERING", "SUSHI", "PIZZA", "FASTFOOD", "CANTEEN", "DONER",
    "COFFEE", "BUBBLETEA",
    "GROCERY", "FRUITSVEGS", "BUILDMAT", "AUTOPARTS",
    "OPTICS", "FLOWERS", "PETSHOP",
    "FURNITURE", "LOFTFURNITURE",
}

# Спец-вопросы: какие применяются к каким нишам
SPECIFIC_QUESTIONS_MAP = {
    "Q_NOT_LIC":      ["NOTARY"],
    "Q_IPO":          ["EVALUATION"],
    "Q_MED_LIC":      ["DENTAL", "COSMETOLOGY", "PHARMACY"],
    "Q_EDU_LIC":      ["KIDSCENTER", "LANGUAGES", "KINDERGARTEN"],
    "Q_BANK_ACCR":    ["EVALUATION", "REALTOR"],
    "Q_FRANCHISE_PVZ":["PVZ"],
    "Q_CHAIRS":       ["BARBER", "MANICURE", "BROW", "LASH", "SUGARING", "MASSAGE",
                       "BEAUTY", "DENTAL", "COSMETOLOGY"],
    "Q_POSTS":        ["AUTOSERVICE", "TIRESERVICE", "CARWASH", "DETAILING"],
    "Q_KITCHEN":      ["PIZZA", "SUSHI", "COFFEE", "BUBBLETEA", "FASTFOOD", "DONER", "CANTEEN"],
    "Q_KIDS_FMT":     ["KIDSCENTER"],
    "Q_PHOTO_FMT":    ["PHOTO"],
    "Q_LANG_FMT":     ["LANGUAGES"],
    "Q_HOTEL_CLASS":  ["HOTEL"],
    "Q_MANIC_SPEC":   ["MANICURE"],
    "Q_BARBER_TYPE":  ["BARBER"],
    "Q_CARS_BRAND":   ["AUTOSERVICE", "TIRESERVICE", "DETAILING"],
    "Q_GROCERY_FMT":  ["GROCERY"],
}

SPECIFIC_QUESTIONS_CATALOG = [
    ("Q_NOT_LIC",       "Вы уже имеете лицензию нотариуса?",
                         "да|нет|в процессе"),
    ("Q_IPO",           "Вы уже имеете лицензию оценщика?",
                         "да|нет|в процессе"),
    ("Q_MED_LIC",       "Есть ли у вас медицинская лицензия или специалист с лицензией?",
                         "есть|планирую получить|нет"),
    ("Q_EDU_LIC",       "Планируете официальное образование с выдачей документа?",
                         "да|нет"),
    ("Q_BANK_ACCR",     "Планируете аккредитацию в банках?",
                         "да|нет|позже"),
    ("Q_FRANCHISE_PVZ", "Какие маркетплейсы планируете обслуживать?",
                         "Wildberries|Kaspi Доставка|Ozon|несколько"),
    ("Q_CHAIRS",        "Сколько рабочих мест (кресел/столов)?",
                         "1|2|3-5|6-10|10+"),
    ("Q_POSTS",         "Сколько постов/боксов?",
                         "1|2|3|4+"),
    ("Q_KITCHEN",       "Формат кухни?",
                         "только навынос|с залом|полная кухня-ресторан"),
    ("Q_KIDS_FMT",      "Какой формат развивающего центра?",
                         "раннее развитие|досадиковая группа|подготовка к школе|логопед|Lego/робототехника|игровая в ТЦ"),
    ("Q_PHOTO_FMT",     "Какой формат фотостудии?",
                         "только съёмка|студия с интерьерами|выездная|школьная"),
    ("Q_LANG_FMT",      "Формат языковой школы?",
                         "индивидуально из дома|мини-группы в офисе|полноценная школа"),
    ("Q_HOTEL_CLASS",   "Класс отеля?",
                         "хостел|2-3 звезды|4-5 звёзд|апарт-отель"),
    ("Q_MANIC_SPEC",    "Есть ли специализация?",
                         "базовый маникюр|наращивание|педикюр|всё вместе"),
    ("Q_BARBER_TYPE",   "Формат?",
                         "барбершоп мужской|унисекс|эконом парикмахерская"),
    ("Q_CARS_BRAND",    "Специализация по маркам?",
                         "универсальный|немецкие|японские|премиум"),
    ("Q_GROCERY_FMT",   "Формат точки?",
                         "киоск|минимаркет|супермаркет|нишевый"),
]


def requires_license(nid: str) -> tuple[str, str]:
    if nid in LICENSE_MANDATORY:
        return "mandatory", LICENSE_MANDATORY[nid]
    if nid in LICENSE_OPTIONAL:
        return "optional", LICENSE_OPTIONAL[nid]
    return "no", ""


def self_operation(nid: str) -> str:
    return "yes" if nid in SELF_OP_YES else "no"


def class_grades(nid: str) -> str:
    return "yes" if nid in CLASS_GRADES_YES else "no"


def locations(nid: str) -> tuple[str, str]:
    return LOCATION_MAP.get(nid, ("street,own_building", "street"))


def area_mode(nid: str) -> str:
    if nid in AREA_HIDDEN_IF_HOME:
        return "hidden_if_home"
    return "required"


def staff_mode(nid: str) -> str:
    if nid in STAFF_HIDDEN_MANDATORY_TEAM:
        return "hidden"
    if nid in SELF_OP_YES and nid in CLASS_GRADES_YES:
        # владелец может сам работать, но при крупном формате — нанимает
        return "choice"
    if nid in SELF_OP_YES:
        return "choice"
    return "hired_only"


def specific_questions_for(nid: str) -> str:
    ids = [qid for qid, niches in SPECIFIC_QUESTIONS_MAP.items() if nid in niches]
    return ",".join(ids)


def niche_notes(nid: str) -> str:
    if nid in LICENSE_MANDATORY:
        return "Лицензируемая профессия. Без лицензии запуск невозможен."
    if nid in STAFF_HIDDEN_MANDATORY_TEAM:
        return "Требуется команда с первого дня."
    if nid in AREA_HIDDEN_IF_HOME:
        return "Возможен старт из дома — площадь скрывается если выбран формат home."
    return ""


def build_niches_rows():
    rows = []
    for nid in TARGET_NICHES:
        req_lic, lic_desc = requires_license(nid)
        allowed, default = locations(nid)
        rows.append({
            "niche_id": nid,
            "niche_name": NICHE_NAMES.get(nid, nid),
            "requires_license": req_lic,
            "license_description": lic_desc,
            "self_operation_possible": self_operation(nid),
            "class_grades_applicable": class_grades(nid),
            "allowed_location_types": allowed,
            "default_location_type": default,
            "area_question_mode": area_mode(nid),
            "staff_question_mode": staff_mode(nid),
            "specific_questions_ids": specific_questions_for(nid),
            "niche_notes": niche_notes(nid),
        })
    return rows


def build_specific_questions_rows():
    out = []
    for qid, text, options in SPECIFIC_QUESTIONS_CATALOG:
        applies = ",".join(SPECIFIC_QUESTIONS_MAP.get(qid, []))
        out.append({
            "question_id": qid,
            "question_text": text,
            "options": options,           # pipe-separated
            "applies_to_niches": applies, # comma-separated
        })
    return out


def write_workbook(out_path: str) -> None:
    wb = Workbook()

    # -------------- Лист «Ниши» ---------------------------------------------
    ws = wb.active
    ws.title = "Ниши"

    # Headers live on Excel row 6 → pandas header=5 (0-indexed).
    # Rows 1-5 are free-form documentation.
    title_font = Font(bold=True, size=14)
    ws.cell(row=1, column=1, value="ZEREK — Ниши: адаптивная анкета v3").font = title_font
    ws.cell(row=2, column=1, value="Каждая строка = одна ниша. Колонки описывают логику вопросов Quick Check v2.")
    ws.cell(row=3, column=1, value=f"Ниш в реестре: {len(TARGET_NICHES)}")
    ws.cell(row=4, column=1, value="Справочник опций — см. листы «Специфичные вопросы» и «Типы локации».")
    # row 5 empty separator

    headers = [
        "niche_id", "niche_name", "requires_license", "license_description",
        "self_operation_possible", "class_grades_applicable",
        "allowed_location_types", "default_location_type",
        "area_question_mode", "staff_question_mode",
        "specific_questions_ids", "niche_notes",
    ]

    header_fill = PatternFill(start_color="1F1F2E", end_color="1F1F2E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=6, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill

    for i, row in enumerate(build_niches_rows(), start=7):
        for j, h in enumerate(headers, start=1):
            ws.cell(row=i, column=j, value=row.get(h, ""))

    # column widths
    widths = {1: 16, 2: 32, 3: 16, 4: 48, 5: 18, 6: 18, 7: 44, 8: 20, 9: 20, 10: 22, 11: 36, 12: 60}
    for col, w in widths.items():
        ws.column_dimensions[ws.cell(row=6, column=col).column_letter].width = w

    # -------------- Лист «Специфичные вопросы» ------------------------------
    ws2 = wb.create_sheet("Специфичные вопросы")
    ws2.cell(row=1, column=1, value="ZEREK — Специфичные вопросы анкеты v3").font = title_font
    ws2.cell(row=2, column=1, value="Вопросы, которые применяются только к определённым нишам.")
    ws2.cell(row=3, column=1, value="options: варианты через |  ·  applies_to_niches: niche_id через запятую")

    sq_headers = ["question_id", "question_text", "options", "applies_to_niches"]
    for col_idx, h in enumerate(sq_headers, start=1):
        c = ws2.cell(row=6, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill

    for i, row in enumerate(build_specific_questions_rows(), start=7):
        for j, h in enumerate(sq_headers, start=1):
            ws2.cell(row=i, column=j, value=row.get(h, ""))

    widths2 = {1: 20, 2: 56, 3: 58, 4: 52}
    for col, w in widths2.items():
        ws2.column_dimensions[ws2.cell(row=6, column=col).column_letter].width = w

    # -------------- Лист-справочник «Типы локации» --------------------------
    ws3 = wb.create_sheet("Типы локации")
    ws3.cell(row=1, column=1, value="Справочник allowed_location_types — для чтения, не править в расчётах.").font = title_font
    loc_rows = [
        ("tc",                  "🏬", "Торговый центр"),
        ("street",              "🏪", "Улица / отдельное помещение"),
        ("home",                "🏠", "Работа из дома"),
        ("highway",             "🛣️", "Возле дороги"),
        ("residential_complex", "🏢", "Коммерция в ЖК"),
        ("business_center",     "🏢", "Бизнес-центр"),
        ("market",              "🛍️", "Рынок / павильон"),
        ("online",              "🌐", "Только онлайн"),
        ("residential_area",    "🏘️", "Спальный район"),
        ("own_building",        "🏛️", "Отдельно стоящее здание"),
    ]
    loc_headers = ["location_id", "icon", "label"]
    for col_idx, h in enumerate(loc_headers, start=1):
        c = ws3.cell(row=6, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
    for i, (lid, ic, lb) in enumerate(loc_rows, start=7):
        ws3.cell(row=i, column=1, value=lid)
        ws3.cell(row=i, column=2, value=ic)
        ws3.cell(row=i, column=3, value=lb)
    for col, w in {1: 22, 2: 6, 3: 32}.items():
        ws3.column_dimensions[ws3.cell(row=6, column=col).column_letter].width = w

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"✅ {out_path} — {len(TARGET_NICHES)} ниш, {len(SPECIFIC_QUESTIONS_CATALOG)} спец-вопросов, {len(loc_rows)} типов локации")


if __name__ == "__main__":
    repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out = os.path.join(repo_root, "data", "kz", "07_niches.xlsx")
    write_workbook(out)
