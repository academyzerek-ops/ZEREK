"""
Builds data/kz/08_niche_formats.xlsx — fallback format catalog with v1.0 spec fields:
  typical_staff, allowed_locations, format_type (SOLO/HOME/MOBILE/KIOSK/HIGHWAY/PRODUCTION/STANDARD).

Format entries describe startup-level businesses only (per spec).
Removed: HOTEL_3STAR, HOTEL_4STAR, TAILOR_FULL.
Added: 7 SOLO formats (BEAUTY_SOLO / NAIL_SOLO / BROW_SOLO / LASH_SOLO / SUGAR_SOLO / MASSAGE_SOLO / COSM_SOLO).

Run: python3 scripts/build_08_niche_formats.py
"""
from __future__ import annotations

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


# ---------------------------------------------------------------------------
# FORMATS — (format_id, format_name, area_m2, capex_standard, class, format_type,
#            allowed_locations|'*'|'auto', typical_staff)
# class: "эконом" / "стандарт" / "премиум" / ""
# format_type: STANDARD | SOLO | HOME | MOBILE | KIOSK | HIGHWAY | PRODUCTION
# allowed_locations: comma-separated list or '*' (любая из 12) or 'auto' (скрыт)
# typical_staff: "роль:кол-во|роль2:кол-во2" или пусто для SOLO
# ---------------------------------------------------------------------------

FORMATS = {
    # ---------- АРХЕТИП A — Услуги с мастерами ----------
    "BARBER": [
        ("BARBER_SOLO",     "Барбер-одиночка",         15,  1800000,  "эконом",    "SOLO",     "rent_in_salon", ""),
        ("BARBER_STANDARD", "Барбершоп 3-5 кресел",    45,  7500000,  "стандарт",  "STANDARD", "city_center,residential,residential_complex,mall_standard", "барбер:4|администратор:1"),
        ("BARBER_PREMIUM",  "Премиум барбершоп",       80, 18000000,  "премиум",   "STANDARD", "city_center,mall_premium,bc_premium", "барбер:5|администратор:1|бариста:1"),
    ],
    "BEAUTY": [
        ("BEAUTY_SOLO",     "Мастер в чужом салоне",    0,   400000,  "эконом",    "SOLO", "rent_in_salon", ""),
        ("BEAUTY_STANDARD", "Салон красоты",           80, 16000000,  "стандарт",  "STANDARD", "city_center,residential,residential_complex,mall_standard", "парикмахер:2|мастер_маникюра:2|косметолог:1|администратор:1"),
        ("BEAUTY_PREMIUM",  "Премиум салон",          120, 32000000,  "премиум",   "STANDARD", "city_center,mall_premium,bc_premium", "парикмахер:3|мастер_маникюра:2|косметолог:2|администратор:2"),
    ],
    "MANICURE": [
        ("NAIL_HOME",       "Мастер на дому",          10,   350000,  "эконом",    "HOME", "auto", ""),
        ("NAIL_SOLO",       "Мастер в чужом салоне",    0,   300000,  "эконом",    "SOLO", "rent_in_salon", ""),
        ("NAIL_CABINET",    "Кабинет 1-2 мастера",     20,  2200000,  "стандарт",  "STANDARD", "residential,residential_complex,mall_standard", "мастер_маникюра:2"),
        ("NAIL_SALON",      "Студия маникюра",         50,  9000000,  "премиум",   "STANDARD", "city_center,mall_premium,mall_standard", "мастер_маникюра:4|администратор:1"),
    ],
    "BROW": [
        ("BROW_HOME",       "Мастер на дому",           8,   300000,  "эконом",    "HOME", "auto", ""),
        ("BROW_SOLO",       "Мастер в чужом салоне",    0,   250000,  "эконом",    "SOLO", "rent_in_salon", ""),
        ("BROW_CABINET",    "Свой кабинет",            18,  2000000,  "стандарт",  "STANDARD", "residential,residential_complex,mall_standard", "мастер:2"),
    ],
    "LASH": [
        ("LASH_HOME",       "Мастер на дому",           8,   300000,  "эконом",    "HOME", "auto", ""),
        ("LASH_SOLO",       "Мастер в чужом салоне",    0,   250000,  "эконом",    "SOLO", "rent_in_salon", ""),
        ("LASH_CABINET",    "Свой кабинет",            18,  2000000,  "стандарт",  "STANDARD", "residential,residential_complex,mall_standard", "мастер:2"),
    ],
    "SUGARING": [
        ("SUGAR_HOME",      "Мастер на дому",          10,   350000,  "эконом",    "HOME", "auto", ""),
        ("SUGAR_SOLO",      "Мастер в чужом салоне",    0,   300000,  "эконом",    "SOLO", "rent_in_salon", ""),
        ("SUGAR_CABINET",   "Кабинет шугаринга",       20,  2500000,  "стандарт",  "STANDARD", "residential,residential_complex,mall_standard", "мастер:2"),
    ],
    "MASSAGE": [
        ("MASSAGE_HOME",    "Выезд на дом",            10,   300000,  "эконом",    "HOME", "auto", ""),
        ("MASSAGE_SOLO",    "Мастер в чужом салоне",    0,   400000,  "эконом",    "SOLO", "rent_in_salon", ""),
        ("MASSAGE_STUDIO",  "Массажная студия",        40,  7000000,  "стандарт",  "STANDARD", "residential,residential_complex,mall_standard", "массажист:2|администратор:1"),
    ],
    "COSMETOLOGY": [
        ("COSM_SOLO",       "Косметолог в чужом салоне", 0,  800000,  "эконом",    "SOLO", "rent_in_salon", ""),
        ("COSM_CABINET",    "Кабинет косметолога",     25,  6000000,  "стандарт",  "STANDARD", "residential,residential_complex,mall_standard,bc_standard", "косметолог:1|администратор:1"),
        ("COSM_CLINIC",     "Мини-клиника",            60, 22000000,  "премиум",   "STANDARD", "city_center,mall_premium,bc_premium", "косметолог:2|врач:1|администратор:1"),
    ],
    "DENTAL": [
        ("DENTAL_1CH",      "Кабинет 1 кресло",        30, 12000000,  "эконом",    "STANDARD", "residential,residential_complex,bc_standard", "стоматолог:1|ассистент:1|администратор:1"),
        ("DENTAL_CLINIC",   "Клиника 3-5 кресел",      80, 45000000,  "стандарт",  "STANDARD", "city_center,residential,mall_standard", "стоматолог:3|ассистент:3|администратор:1"),
    ],
    "OPTICS": [
        ("OPTICS_TC",       "Оптика в ТЦ",             40, 16000000,  "стандарт",  "STANDARD", "mall_premium,mall_standard", "продавец:2|офтальмолог:1"),
        ("OPTICS_STREET",   "Оптика-улица",            70, 24000000,  "премиум",   "STANDARD", "city_center,residential,residential_complex", "продавец:2|офтальмолог:1|администратор:1"),
    ],
    "TAILOR": [
        ("TAILOR_HOME",     "Ателье на дому",          15,   500000,  "эконом",    "HOME", "auto", ""),
        ("TAILOR_MINI",     "Мини-ателье",             25,  2500000,  "стандарт",  "STANDARD", "residential,residential_complex,mall_standard,market", "закройщик:1|швея:1"),
    ],
    "REPAIR_PHONE": [
        ("REPAIR_POINT",    "Точка в ТЦ",              12,  1800000,  "эконом",    "KIOSK", "mall_premium,mall_standard,bc_standard", "мастер:1"),
        ("REPAIR_STUDIO",   "Мастерская",              25,  4500000,  "стандарт",  "STANDARD", "city_center,residential,mall_standard", "мастер:2"),
    ],
    "PHOTO": [
        ("PHOTO_HOME",      "Выездной фотограф",        0,  1500000,  "эконом",    "MOBILE", "auto", ""),
        ("PHOTO_STUDIO",    "Фотостудия с интерьерами", 60, 10000000, "стандарт",  "STANDARD", "city_center,residential_complex,bc_standard,industrial", "фотограф:1|администратор:1"),
    ],
    "DETAILING": [
        ("DETAIL_BOX1",     "Одиночный бокс",          80,  6500000,  "стандарт",  "HIGHWAY", "auto", "детейлер:2"),
        ("DETAIL_STUDIO",   "Детейлинг-студия",       150, 22000000,  "премиум",   "STANDARD", "city_center,industrial,highway", "детейлер:3|администратор:1"),
    ],
    "NOTARY": [
        ("NOTARY_OFFICE",   "Нотариальная контора",    30,  5500000,  "стандарт",  "STANDARD", "city_center,bc_premium,bc_standard", "нотариус:1|помощник:1"),
    ],

    # ---------- АРХЕТИП B — Общепит ----------
    "COFFEE": [
        ("COFFEE_KIOSK",    "Островок / кофе с собой", 12,  6000000,  "эконом",    "KIOSK", "mall_premium,mall_standard,bc_premium,bc_standard", "бариста:2"),
        ("COFFEE_CAFE",     "Кофейня-кафе",            60, 18000000,  "стандарт",  "STANDARD", "city_center,residential,residential_complex,mall_standard", "бариста:3|кассир:1"),
        ("COFFEE_SPECIAL",  "Specialty coffee",        80, 32000000,  "премиум",   "STANDARD", "city_center,mall_premium,bc_premium", "бариста:3|кассир:1|администратор:1"),
    ],
    "BAKERY": [
        ("BAKERY_SMALL",    "Мини-пекарня с витриной", 45, 12000000,  "эконом",    "STANDARD", "residential,residential_complex,market", "пекарь:2|продавец:1"),
        ("BAKERY_FULL",     "Пекарня-кафе",            80, 22000000,  "стандарт",  "STANDARD", "city_center,residential,residential_complex,mall_standard", "пекарь:3|продавец:2"),
        ("BAKERY_PRODUCTION","Производство + B2B",    120, 32000000,  "премиум",   "PRODUCTION", "auto", "пекарь:4|менеджер_b2b:1|водитель:1"),
    ],
    "CONFECTION": [
        ("CONFECT_HOME",    "Домашний кондитер",       20,  1500000,  "эконом",    "HOME", "auto", ""),
        ("CONFECT_WORKSHOP","Кондитерский цех",        60, 12000000,  "стандарт",  "STANDARD", "residential,residential_complex,industrial", "кондитер:3|помощник:1"),
    ],
    "DONER": [
        ("DONER_TAKEOUT",   "Донерная навынос",        18,  5500000,  "эконом",    "STANDARD", "residential,residential_complex,market", "повар:2|кассир:1"),
        ("DONER_CAFE",      "Донерная с залом",        55, 14000000,  "стандарт",  "STANDARD", "city_center,residential,mall_standard", "повар:3|кассир:1|официант:1"),
    ],
    "FASTFOOD": [
        ("FASTFOOD_TAKE",   "Такаут в ТЦ",             25,  7000000,  "эконом",    "KIOSK", "mall_premium,mall_standard,bc_standard", "повар:2|кассир:1"),
        ("FASTFOOD_CAFE",   "Фастфуд с залом",         70, 18000000,  "стандарт",  "STANDARD", "city_center,residential,mall_standard", "повар:3|кассир:2|официант:1"),
    ],
    "PIZZA": [
        ("PIZZA_DELIVERY",  "Доставка пиццы",          30,  9000000,  "эконом",    "MOBILE", "auto", "пиццмейкер:2|курьер:2"),
        ("PIZZA_CAFE",      "Пиццерия с залом",        80, 22000000,  "стандарт",  "STANDARD", "city_center,residential,mall_standard", "пиццмейкер:3|кассир:1|официант:2"),
    ],
    "SUSHI": [
        ("SUSHI_DELIVERY",  "Доставка суши",           30,  9000000,  "эконом",    "MOBILE", "auto", "сушист:2|курьер:2"),
        ("SUSHI_BAR",       "Суши-бар с залом",        75, 26000000,  "стандарт",  "STANDARD", "city_center,residential,mall_standard", "сушист:3|кассир:1|официант:2"),
    ],
    "BUBBLETEA": [
        ("BUBBLE_KIOSK",    "Островок в ТЦ",           10,  4500000,  "эконом",    "KIOSK", "mall_premium,mall_standard,bc_premium,bc_standard", "бариста:2"),
        ("BUBBLE_CAFE",     "Бабл-ти с залом",         40, 12000000,  "стандарт",  "STANDARD", "city_center,mall_standard,residential_complex", "бариста:3|кассир:1"),
    ],
    "CANTEEN": [
        ("CANTEEN_BC",      "Столовая в БЦ",           90, 14000000,  "стандарт",  "STANDARD", "bc_standard,bc_premium", "повар:3|кассир:2|помощник:2"),
        ("CANTEEN_ENTER",   "Столовая при предприятии",140, 20000000,  "стандарт",  "PRODUCTION", "auto", "повар:4|кассир:2|помощник:2"),
    ],
    "SEMIFOOD": [
        ("SEMI_HOME",       "Домашнее производство",   25,  2500000,  "эконом",    "HOME", "auto", ""),
        ("SEMI_WORKSHOP",   "Мини-цех",                70, 14000000,  "стандарт",  "PRODUCTION", "auto", "повар:3|помощник:2|менеджер:1"),
    ],
    "CATERING": [
        ("CATERING_HOME",   "Домашний кейтеринг",      35,  3000000,  "эконом",    "MOBILE", "auto", ""),
        ("CATERING_STUDIO", "Студия-цех",              80, 15000000,  "стандарт",  "MOBILE", "auto", "повар:3|помощник:2|менеджер:1"),
        ("CATERING_EVENT",  "Event-кейтеринг",        150, 40000000,  "премиум",   "MOBILE", "auto", "повар:5|помощник:4|менеджер:2|логист:1"),
    ],

    # ---------- АРХЕТИП C — Розничная торговля ----------
    "MEATSHOP": [
        ("MEAT_SMALL",      "Небольшая лавка",         25,  5500000,  "эконом",    "STANDARD", "residential,market,residential_complex", "продавец:1|разделщик:1"),
        ("MEAT_STANDARD",   "Мясная лавка",            50, 12000000,  "стандарт",  "STANDARD", "residential,residential_complex,market", "продавец:2|разделщик:1"),
    ],
    "FRUITSVEGS": [
        ("FV_KIOSK",        "Киоск",                   10,  2500000,  "эконом",    "STANDARD", "market,residential,residential_complex", "продавец:1"),
        ("FV_STORE",        "Магазин",                 40,  7000000,  "стандарт",  "STANDARD", "residential,residential_complex,market", "продавец:2"),
    ],
    "FLOWERS": [
        ("FLOWER_KIOSK",    "Цветочный киоск",         12,  2500000,  "эконом",    "STANDARD", "city_center,residential,market,residential_complex", "флорист:1"),
        ("FLOWER_STUDIO",   "Цветочная студия",        40,  8000000,  "стандарт",  "STANDARD", "city_center,residential,residential_complex,mall_standard", "флорист:2|курьер:1"),
    ],
    "AUTOPARTS": [
        ("PARTS_STORE",     "Магазин автозапчастей",   80, 12000000,  "стандарт",  "STANDARD", "residential,market,industrial,highway", "продавец:2|консультант:1"),
    ],
    "BUILDMAT": [
        ("BUILD_SMALL",     "Небольшой магазин",       80, 12000000,  "стандарт",  "STANDARD", "residential,industrial,market", "продавец:2|водитель:1"),
        ("BUILD_FULL",      "Строймаркет",            250, 48000000,  "премиум",   "STANDARD", "industrial,highway,suburb", "продавец:3|водитель:2|менеджер:1"),
    ],
    "PETSHOP": [
        ("PET_SMALL",       "Зоомагазин мини",         30,  5000000,  "эконом",    "STANDARD", "residential,residential_complex,market", "продавец:1"),
        ("PET_FULL",        "Зоомагазин",              80, 14000000,  "стандарт",  "STANDARD", "residential,mall_standard,residential_complex", "продавец:2|грумер:1"),
    ],
    "PHARMACY": [
        ("PHARM_SMALL",     "Небольшая аптека",        40, 10000000,  "эконом",    "STANDARD", "residential,residential_complex,suburb", "фармацевт:2"),
        ("PHARM_FULL",      "Сетевая аптека",          80, 22000000,  "стандарт",  "STANDARD", "city_center,residential,mall_standard", "фармацевт:3|кассир:1"),
    ],
    "GROCERY": [
        ("GROC_KIOSK",      "Минимаркет",              25,  3500000,  "эконом",    "STANDARD", "residential,residential_complex,suburb", "продавец:2"),
        ("GROC_STANDARD",   "Минимаркет",              80, 14000000,  "стандарт",  "STANDARD", "residential,residential_complex,market", "продавец:3|кассир:1"),
        ("GROC_SUPER",      "Супермаркет",            250, 48000000,  "премиум",   "STANDARD", "city_center,residential,mall_standard", "продавец:5|кассир:3|менеджер:1"),
    ],
    "WATERPLANT": [
        ("WATER_MINI",      "Мини-линия розлива",     100, 18000000,  "стандарт",  "PRODUCTION", "auto", "оператор:2|водитель:1|менеджер:1"),
        ("WATER_FULL",      "Завод розлива",          300, 55000000,  "премиум",   "PRODUCTION", "auto", "оператор:4|водитель:2|менеджер:1|логист:1"),
    ],

    # ---------- АРХЕТИП D — Абонементы / подписка ----------
    "FITNESS": [
        ("FIT_STUDIO",      "Студия 100 м²",          100, 14000000,  "эконом",    "STANDARD", "residential,residential_complex,mall_standard", "тренер:2|администратор:1"),
        ("FIT_CLUB",        "Фитнес-клуб 500 м²",     500, 80000000,  "стандарт",  "PRODUCTION", "auto", "тренер:5|администратор:3|уборщик:2"),
    ],
    "YOGA": [
        ("YOGA_STUDIO",     "Йога-студия",             80,  6000000,  "стандарт",  "STANDARD", "residential_complex,city_center,mall_standard", "тренер:2|администратор:1"),
    ],
    "LANGUAGES": [
        ("LANG_HOME",       "Репетитор из дома",       15,   400000,  "эконом",    "HOME", "auto", ""),
        ("LANG_MICRO",      "Мини-школа",              60,  3500000,  "стандарт",  "STANDARD", "residential,residential_complex,bc_standard", "преподаватель:3|администратор:1"),
        ("LANG_FULL",       "Языковая школа",         150, 14000000,  "премиум",   "STANDARD", "city_center,mall_standard,bc_standard", "преподаватель:5|администратор:1|методист:1"),
    ],
    "KIDSCENTER": [
        ("KIDS_STUDIO",     "Развивающая студия",      80,  5000000,  "стандарт",  "STANDARD", "residential,residential_complex,mall_standard", "педагог:3|администратор:1"),
        ("KIDS_LARGE",      "Центр с несколькими программами",200, 16000000, "премиум", "STANDARD", "city_center,residential,mall_standard", "педагог:5|администратор:1|методист:1"),
    ],
    "KINDERGARTEN": [
        ("KG_HOME",         "Домашний сад 8-12 детей", 70,  2500000,  "эконом",    "HOME", "auto", ""),
        ("KG_STANDARD",     "Частный сад 30+ детей",  200, 18000000,  "стандарт",  "PRODUCTION", "auto", "воспитатель:4|повар:1|медсестра:1|администратор:1"),
    ],
    "ACCOUNTING": [
        ("ACC_SOLO",        "Самозанятый бухгалтер",    0,   250000,  "эконом",    "SOLO", "rent_in_salon", ""),
        ("ACC_AGENCY",      "Бухгалтерское агентство", 40,  3500000,  "стандарт",  "STANDARD", "bc_standard,bc_premium,city_center", "бухгалтер:3|менеджер:1"),
    ],
    "CROSSFIT": [
        ("CROSSFIT_STUDIO", "Кроссфит-зал",           200, 18000000,  "стандарт",  "PRODUCTION", "auto", "тренер:3|администратор:1"),
    ],
    "MARTIAL_ARTS": [
        ("MMA_STUDIO",      "Зал единоборств",        150, 10000000,  "стандарт",  "STANDARD", "residential,industrial,mall_standard", "тренер:3|администратор:1"),
    ],
    "FOOTBALL_SCHOOL": [
        ("FB_SMALL",        "Детская футбольная школа",  0, 3000000,  "эконом",    "MOBILE", "auto", "тренер:3|администратор:1"),
    ],
    "GROUP_FITNESS": [
        ("GF_STUDIO",       "Студия групповых",       100,  8000000,  "стандарт",  "STANDARD", "residential_complex,mall_standard,city_center", "тренер:3|администратор:1"),
    ],

    # ---------- АРХЕТИП E — Проектный / заказной ----------
    "REALTOR": [
        ("REAL_SOLO",       "Независимый риэлтор",      0,   200000,  "эконом",    "SOLO", "rent_in_salon", ""),
        ("REAL_AGENCY",     "Агентство недвижимости",  50,  4500000,  "стандарт",  "STANDARD", "bc_standard,bc_premium,city_center", "риэлтор:3|менеджер:1"),
    ],
    "EVALUATION": [
        ("EVAL_SOLO",       "Оценщик-самозанятый",      0,   400000,  "эконом",    "SOLO", "rent_in_salon", ""),
        ("EVAL_AGENCY",     "Оценочная компания",      30,  3000000,  "стандарт",  "STANDARD", "bc_standard,bc_premium,city_center", "оценщик:2|помощник:1"),
    ],
    "FURNITURE": [
        ("FURN_SMALL",      "Цех мебели на заказ",    120, 14000000,  "стандарт",  "PRODUCTION", "auto", "столяр:3|монтажник:1|дизайнер:1"),
        ("FURN_FULL",       "Мебельное производство", 300, 42000000,  "премиум",   "PRODUCTION", "auto", "столяр:5|монтажник:2|дизайнер:1|менеджер:1"),
    ],
    "LOFTFURNITURE": [
        ("LOFT_SMALL",      "Мастерская лофт-мебели", 120, 12000000,  "стандарт",  "PRODUCTION", "auto", "сварщик:2|столяр:1|монтажник:1"),
        ("LOFT_FULL",       "Цех металлоконструкций", 250, 32000000,  "премиум",   "PRODUCTION", "auto", "сварщик:3|столяр:1|монтажник:2|менеджер:1"),
    ],
    "PRINTING": [
        ("PRINT_SMALL",     "Мини-типография",         40,  6500000,  "стандарт",  "STANDARD", "city_center,industrial,bc_standard", "оператор:2|менеджер:1"),
        ("PRINT_FULL",      "Полноценная типография",150, 38000000,  "премиум",   "PRODUCTION", "auto", "оператор:3|менеджер:1|дизайнер:1|водитель:1"),
    ],
    "CARGO": [
        ("CARGO_SOLO",      "1 машина (InDrive)",       0,  8000000,  "эконом",    "MOBILE", "auto", ""),
        ("CARGO_FLEET",     "Парк 5+ машин",            0, 45000000,  "стандарт",  "MOBILE", "auto", "водитель:5|диспетчер:1|механик:1"),
    ],

    # ---------- АРХЕТИП F — Мощность / пропускная ----------
    "AUTOSERVICE": [
        ("AUTOSERV_1POST",  "Автосервис 1-2 поста",    80,  5000000,  "эконом",    "HIGHWAY", "auto", "автослесарь:2"),
        ("AUTOSERV_FULL",   "СТО 3-6 постов",         200, 20000000,  "стандарт",  "HIGHWAY", "auto", "автослесарь:4|приёмщик:1|менеджер:1"),
    ],
    "TIRESERVICE": [
        ("TIRE_SMALL",      "Шиномонтаж 1 пост",       40,  3000000,  "эконом",    "HIGHWAY", "auto", "шиномонтажник:2"),
        ("TIRE_STANDARD",   "Шиномонтаж 2-3 поста",    80,  8000000,  "стандарт",  "HIGHWAY", "auto", "шиномонтажник:3|приёмщик:1"),
    ],
    "CARWASH": [
        ("WASH_SELF",       "Самомойка",              250, 28000000,  "эконом",    "HIGHWAY", "auto", "оператор:1"),
        ("WASH_MANUAL",     "Ручная мойка",           120, 14000000,  "стандарт",  "HIGHWAY", "auto", "мойщик:3|администратор:1"),
    ],
    "CLEAN": [
        ("CLEAN_SOLO",      "Клининг-ИП",              10,   500000,  "эконом",    "SOLO", "rent_in_salon", ""),
        ("CLEAN_TEAM",      "Клининговая бригада",     60,  4000000,  "стандарт",  "MOBILE", "auto", "уборщик:4|менеджер:1"),
    ],
    "CARPETCLEAN": [
        ("CARPET_WORKSHOP", "Цех чистки ковров",      120, 10000000,  "стандарт",  "MOBILE", "auto", "оператор:2|водитель:1"),
    ],
    "DRYCLEAN": [
        ("DRY_APPR",        "Приёмка / пункт",         20,  2500000,  "эконом",    "STANDARD", "residential,residential_complex,mall_standard", "приёмщик:1"),
        ("DRY_FULL",        "Химчистка с оборудованием", 80, 18000000, "стандарт", "PRODUCTION", "auto", "оператор:2|приёмщик:1|менеджер:1"),
    ],
    "LAUNDRY": [
        ("LAUNDRY_SELF",    "Прачечная самообслуживания", 60, 14000000, "стандарт", "STANDARD", "residential_complex,bc_standard,market", "оператор:1"),
        ("LAUNDRY_FULL",    "Промышленная прачечная",  200, 42000000,  "премиум",   "PRODUCTION", "auto", "оператор:3|водитель:1|менеджер:1"),
    ],
    "COMPCLUB": [
        ("CC_SMALL",        "Клуб 10-15 мест",         80, 18000000,  "стандарт",  "STANDARD", "residential,residential_complex,mall_standard", "администратор:2"),
        ("CC_LARGE",        "Киберарена 30+ мест",    200, 60000000,  "премиум",   "PRODUCTION", "auto", "администратор:3|техник:1|менеджер:1"),
    ],
    "DRIVING": [
        ("DRIVE_SMALL",     "Автошкола 2-3 машины",    50, 12000000,  "эконом",    "STANDARD", "residential,bc_standard,city_center", "инструктор:2|администратор:1"),
        ("DRIVE_FULL",      "Автошкола 8+ машин",     120, 32000000,  "стандарт",  "STANDARD", "residential,city_center,suburb", "инструктор:5|администратор:1|методист:1"),
    ],
    "HOTEL": [
        ("HOTEL_HOSTEL",    "Хостел 20 койко-мест",   150, 18000000,  "эконом",    "PRODUCTION", "auto", "администратор:2|горничная:2"),
    ],
    "PVZ": [
        ("PVZ_SMALL",       "ПВЗ 15 м²",               15,  2500000,  "эконом",    "STANDARD", "residential,residential_complex,market", "оператор:1"),
        ("PVZ_FULL",        "ПВЗ 35 м²",               35,  4500000,  "стандарт",  "STANDARD", "residential,residential_complex,bc_standard", "оператор:2"),
    ],
}


def build_rows():
    out = []
    for niche_id, lst in FORMATS.items():
        for (fid, fname, area, capex, cls, ftype, allowed_locs, staff) in lst:
            out.append({
                "niche_id":          niche_id,
                "format_id":         fid,
                "format_name":       fname,
                "area_m2":           area,
                "capex_standard":    capex,
                "class":             cls,
                "format_type":       ftype,
                "allowed_locations": allowed_locs,
                "typical_staff":     staff,
            })
    return out


def write_workbook(out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Форматы"

    title_font = Font(bold=True, size=14)
    ws.cell(row=1, column=1, value="ZEREK — Форматы ниш (v1.0 spec)").font = title_font
    ws.cell(row=2, column=1, value="format_type ∈ SOLO|HOME|MOBILE|KIOSK|HIGHWAY|PRODUCTION|STANDARD")
    ws.cell(row=3, column=1, value="allowed_locations: CSV из config/locations.yaml или 'auto' (скрыт)")
    ws.cell(row=4, column=1, value="typical_staff: формат 'роль:n|роль2:m' или пусто для SOLO")

    headers = ["niche_id", "format_id", "format_name", "area_m2", "capex_standard",
               "class", "format_type", "allowed_locations", "typical_staff"]
    header_fill = PatternFill(start_color="1F1F2E", end_color="1F1F2E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=6, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill

    rows = build_rows()
    for i, row in enumerate(rows, start=7):
        for j, h in enumerate(headers, start=1):
            ws.cell(row=i, column=j, value=row.get(h, ""))

    widths = {1: 16, 2: 22, 3: 36, 4: 10, 5: 18, 6: 12, 7: 14, 8: 60, 9: 54}
    for col, w in widths.items():
        ws.column_dimensions[ws.cell(row=6, column=col).column_letter].width = w

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"✅ {out_path}")
    print(f"   Ниш: {len(FORMATS)}")
    print(f"   Форматов: {len(rows)}")
    # Статистика по format_type
    type_counts = {}
    for r in rows:
        t = r["format_type"]
        type_counts[t] = type_counts.get(t, 0) + 1
    print(f"   По типам: {type_counts}")


if __name__ == "__main__":
    repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out = os.path.join(repo_root, "data", "kz", "08_niche_formats.xlsx")
    write_workbook(out)
