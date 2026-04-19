"""
Builds data/kz/09_surveys.xlsx — source of truth for adaptive surveys.

Three sheets, header on row 6 (pandas header=5):
  1. «Вопросы»      — catalog of every qid with type/options/metadata
  2. «Применимость» — qid × niche_id × tier (express|finmodel) × order
  3. «Зависимости»  — optional conditional-show logic

Run: python3 scripts/build_09_surveys.py
"""
from __future__ import annotations

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


# ---------------------------------------------------------------------------
# QUESTIONS CATALOG — tuples of (qid, question_text, input_type, options|"",
#                                placeholder|"", min, max, step, unit, help)
# input_type ∈ {select_remote, select_format, select, radio, radio_h, number,
#              range, boolean, multi_select}
# options: varianty через "|" (для select/radio/multi_select)
# ---------------------------------------------------------------------------

Q = [
    # ---------- UNIVERSAL ----------
    ("U_CITY",          "Город",                          "select_remote",   "cities",                                   "",          0,0,0, "",    "15 городов РК"),
    ("U_NICHE",         "Ниша",                           "select_remote",   "niches",                                   "",          0,0,0, "",    "Выбирается до анкеты"),
    ("U_CAPITAL",       "Стартовый капитал",              "range",           "",                                         "5000000",   500000, 100000000, 500000, "₸", "Включая кредит"),
    ("U_START_MONTH",   "Месяц старта",                   "select",          "январь|февраль|март|апрель|май|июнь|июль|август|сентябрь|октябрь|ноябрь|декабрь", "", 1,12,1, "", "Влияет на сезонность"),
    ("U_ENTITY",        "Юр.лицо",                        "radio",           "Самозанятый|ИП|ТОО",                       "",          0,0,0, "",    ""),
    ("U_CREDIT",        "Берёте кредит?",                 "boolean",         "",                                         "",          0,0,0, "",    ""),
    ("U_CREDIT_AMOUNT", "Сумма кредита",                  "range",           "",                                         "2000000",   500000, 50000000, 500000, "₸", ""),
    ("U_CREDIT_RATE",   "Ставка (% годовых)",             "range",           "",                                         "22",        12, 30, 1, "%", ""),
    ("U_CREDIT_TERM",   "Срок кредита",                   "range",           "",                                         "36",        12, 60, 6, "мес", ""),
    ("U_GROWTH_Y1",     "Целевой рост выручки за год",    "radio",           "0%|15%|30%|50%|100%",                      "",          0,0,0, "%",   "Консерв / умерен / амбиц"),

    # ---------- ARCHETYPE A (beauty-singles) ----------
    ("A_FORMAT",        "Формат салона",                  "select_format",   "",                                         "",          0,0,0, "",    ""),
    ("A_LOC",           "Локация",                        "select",          "street|tc|residential_complex|home",       "",          0,0,0, "",    "Если формат не home"),
    ("A_AREA",          "Площадь (м²)",                   "number",          "",                                         "25",        5, 500, 5, "м²", ""),
    ("A_CHAIRS",        "Сколько рабочих мест",           "radio",           "1|2|3-5|6-10|10+",                         "",          0,0,0, "",    ""),
    ("A_STAFF_MODE",    "Кто работает",                   "radio",           "Сам работаю|Нанимаю мастеров",             "",          0,0,0, "",    ""),
    ("A_CHECK",         "Средний чек (₸)",                "number",          "",                                         "3500",      500, 50000, 100, "₸", ""),
    ("A_LOAD",          "Клиентов на одно кресло в день", "radio",           "1|2|3|4-5|6-8",                            "",          0,0,0, "",    ""),
    ("A_COGS",          "Доля расходников от чека",       "radio",           "5%|10%|15%|20%",                           "",          0,0,0, "%",   "Краски/расходка"),
    ("A_MARKETING",     "Маркетинг (₸/мес)",              "number",          "",                                         "80000",     0, 2000000, 10000, "₸", ""),
    ("A_FOT_MASTER",    "Оплата мастера",                 "radio",           "40% от чека|50% от чека|60% от чека|фикс ставка", "", 0,0,0, "", ""),
    ("A_RENT",          "Аренда (₸/мес)",                 "number",          "",                                         "150000",    0, 3000000, 10000, "₸", ""),

    # ---------- ARCHETYPE B (общепит) ----------
    ("O_FORMAT",        "Формат кухни",                   "radio",           "только навынос|с залом|полная кухня-ресторан|островок",    "", 0,0,0, "", ""),
    ("O_LOC",           "Локация",                        "select",          "tc|street|residential_complex|business_center|market",     "", 0,0,0, "", ""),
    ("O_AREA",          "Площадь (м²)",                   "number",          "",                                         "60",        10, 1000, 5, "м²", ""),
    ("O_CHECK",         "Средний чек (₸)",                "number",          "",                                         "2500",      500, 30000, 100, "₸", ""),
    ("O_TRAFFIC",       "Трафик/день (клиентов)",         "radio",           "30|60|100|200+",                           "",          0,0,0, "", ""),
    ("O_DELIVERY_SHARE","Доля доставки (%)",              "radio",           "0%|20%|40%|60%",                           "",          0,0,0, "%", "Свои или агрегаторы"),
    ("O_STAFF_COUNT",   "Команда",                        "radio",           "2|3-4|5-7|8+",                             "",          0,0,0, "чел", ""),
    ("O_FOODCOST",      "Food cost (% от чека)",          "radio",           "25%|30%|35%|40%",                          "",          0,0,0, "%", "Дефолт по нише"),
    ("O_AGGREGATOR_FEE","Комиссия агрегаторов",           "radio",           "15%|20%|25%|30%",                          "",          0,0,0, "%", "Только если доставка > 0"),
    ("O_RENT_MODE",     "Аренда",                         "radio",           "Фикс ₸/мес|% от выручки",                  "",          0,0,0, "", ""),
    ("O_RENT_VAL",      "Сумма аренды",                   "number",          "",                                         "400000",    0, 5000000, 10000, "₸/мес", ""),
    ("O_MARKETING",     "Маркетинг (₸/мес)",              "number",          "",                                         "150000",    0, 3000000, 10000, "₸", ""),
    ("O_SEASON",        "Сезонность",                     "radio",           "Ровно|Летом ↑|Зимой ↑|Двойной пик",        "",          0,0,0, "", ""),
    ("O_HOURS",         "Часы работы",                    "radio",           "8|12|16|24",                               "",          0,0,0, "ч", ""),

    # ---------- ARCHETYPE C (B2B-prof) ----------
    ("P_FORMAT",        "Формат",                         "radio",           "Соло из дома|Соло офис|Агентство|Сетевое",  "",         0,0,0, "", ""),
    ("P_LICENSE",       "Лицензия / аккредитация",        "radio",           "Есть|В процессе|Нет",                      "",          0,0,0, "",    "Критично для NOTARY/EVAL"),
    ("P_STAFF",         "Кто работает",                   "radio",           "Сам|С помощниками|Команда 3+|Крупное агентство", "",    0,0,0, "", ""),
    ("P_CHECK_OR_RATE", "Средний чек за услугу ИЛИ ставка ₸/час", "number",  "",                                         "25000",     500, 500000, 500, "₸", ""),
    ("P_CLIENTS_PER_MONTH", "Клиентов/услуг в месяц",     "radio",           "до 10|10-30|30-60|60-120|120+",            "",          0,0,0, "", ""),
    ("P_RENT",          "Аренда офиса (₸/мес)",           "number",          "",                                         "100000",    0, 2000000, 10000, "₸", "0 если из дома"),
    ("P_MARKETING",     "Маркетинг (₸/мес)",              "number",          "",                                         "50000",     0, 1000000, 5000, "₸", ""),
    ("P_FOT_SPEC",      "ФОТ специалиста (₸/мес)",        "number",          "",                                         "350000",    0, 2000000, 10000, "₸", "Если есть помощник"),
    ("P_REPEAT_RATE",   "Доля повторных клиентов",        "radio",           "до 20%|20-40%|40-60%|60%+",                "",          0,0,0, "%",  ""),

    # ---------- ARCHETYPE D (auto-service) ----------
    ("D_FORMAT",        "Формат автосервиса",             "radio",           "1 пост|2-3 поста|4+ постов|Премиум-бокс",  "",          0,0,0, "", ""),
    ("D_AREA",          "Площадь (м²)",                   "number",          "",                                         "120",       30, 2000, 10, "м²", ""),
    ("D_POSTS",         "Постов/боксов",                  "radio",           "1|2|3|4+",                                 "",          0,0,0, "", ""),
    ("D_CHECK",         "Средний чек (₸)",                "number",          "",                                         "12000",     1000, 500000, 500, "₸", ""),
    ("D_LOAD",          "Машин на пост в день",           "radio",           "3|5|8|12",                                 "",          0,0,0, "", ""),
    ("D_STAFF_COUNT",   "Команда",                        "radio",           "1|2-3|4-6|7+",                             "",          0,0,0, "чел", ""),
    ("D_BRANDS",        "Специализация",                  "radio",           "Универсальный|Немецкие|Японские|Премиум",  "",          0,0,0, "", ""),
    ("D_COGS",          "Расходники (масла/химия)",       "radio",           "5%|10%|15%",                               "",          0,0,0, "%", ""),
    ("D_PARTS_MARKUP",  "Наценка на запчасти",            "radio",           "0%|10%|20%|30%",                           "",          0,0,0, "%", "Только для AUTOSERVICE"),
    ("D_RENT",          "Аренда (₸/мес)",                 "number",          "",                                         "400000",    0, 5000000, 10000, "₸", ""),
    ("D_MARKETING",     "Маркетинг (₸/мес)",              "number",          "",                                         "100000",    0, 1000000, 5000, "₸", ""),
    ("D_SEASON",        "Сезонность",                     "radio",           "Ровно|Зимой ↑|Летом ↑",                    "",          0,0,0, "", ""),

    # ---------- ARCHETYPE E (retail) ----------
    ("E_FORMAT",        "Формат магазина",                "radio",           "Киоск|Минимаркет|Магазин|Супермаркет|Нишевый",  "",     0,0,0, "", ""),
    ("E_LOC",           "Локация",                        "select",          "residential_area|street|tc|market|residential_complex",  "", 0,0,0, "", ""),
    ("E_AREA",          "Площадь (м²)",                   "number",          "",                                         "50",        10, 2000, 5, "м²", ""),
    ("E_TRAFFIC",       "Трафик/день",                    "radio",           "30|60|100|200+",                           "",          0,0,0, "", ""),
    ("E_CHECK",         "Средний чек (₸)",                "number",          "",                                         "3500",      300, 100000, 100, "₸", ""),
    ("E_MARGIN",        "Средняя наценка",                "radio",           "20%|35%|50%|100%|150%",                    "",          0,0,0, "%", ""),
    ("E_STAFF_COUNT",   "Продавцы",                       "radio",           "1|2-3|4-6|7+",                             "",          0,0,0, "чел", ""),
    ("E_INVENTORY_TURN","Оборачиваемость склада (дней)",  "radio",           "30|45|60|90",                              "",          0,0,0, "дн", ""),
    ("E_LOSS_PCT",      "Доля списаний (скоропорт)",      "radio",           "1%|3%|5%|10%",                             "",          0,0,0, "%", ""),
    ("E_INITIAL_STOCK", "Стартовый закуп товара",         "number",          "",                                         "3000000",   100000, 50000000, 100000, "₸", ""),
    ("E_RENT",          "Аренда (₸/мес)",                 "number",          "",                                         "200000",    0, 5000000, 10000, "₸", ""),
    ("E_MARKETING",     "Маркетинг (₸/мес)",              "number",          "",                                         "80000",     0, 2000000, 5000, "₸", ""),
    ("E_HOURS",         "Часы работы",                    "radio",           "10|12|14|24",                              "",          0,0,0, "ч", ""),

    # ---------- ARCHETYPE F (production) ----------
    ("F_FORMAT",        "Формат производства",            "radio",           "Домашний|Мини-цех|Полноценный цех|Крупное производство", "", 0,0,0, "", ""),
    ("F_AREA",          "Площадь цеха (м²)",              "number",          "",                                         "100",       0, 3000, 10, "м²", "0 если из дома"),
    ("F_CHECK_OR_UNIT", "Средний чек заказа / цена единицы", "number",       "",                                         "80000",     500, 5000000, 500, "₸", ""),
    ("F_VOLUME",        "Заказов/единиц в месяц",         "radio",           "10|30|100|300|1000+",                      "",          0,0,0, "", ""),
    ("F_COGS",          "Доля сырья/материалов",          "radio",           "30%|40%|50%|60%",                          "",          0,0,0, "%", ""),
    ("F_STAFF_COUNT",   "Рабочие в цеху",                 "radio",           "0 (сам)|1-2|3-5|6-10|10+",                 "",          0,0,0, "чел", ""),
    ("F_CLIENT_TYPE",   "Основной клиент",                "radio",           "B2C соцсети|B2B сети|Госзаказ|Смесь",      "",          0,0,0, "", ""),
    ("F_LEAD_TIME",     "Срок производства заказа (дней)","radio",           "1|5|14|30",                                "",          0,0,0, "дн", ""),
    ("F_PREPAY_PCT",    "Предоплата от клиента",          "radio",           "0%|30%|50%|100%",                          "",          0,0,0, "%", ""),
    ("F_EQUIPMENT_CAPEX","Стоимость оборудования",        "number",          "",                                         "5000000",   100000, 100000000, 100000, "₸", ""),
    ("F_RENT",          "Аренда (₸/мес)",                 "number",          "",                                         "200000",    0, 5000000, 10000, "₸", ""),
    ("F_MARKETING",     "Маркетинг + поиск B2B (₸/мес)",  "number",          "",                                         "100000",    0, 2000000, 5000, "₸", ""),
    ("F_LOGISTICS",     "Доля логистики",                 "radio",           "0%|5%|10%|15%",                            "",          0,0,0, "%", ""),

    # ---------- ARCHETYPE G (education/kids) ----------
    ("G_FORMAT",        "Формат",                         "radio",           "Домашний|Небольшой|Полноценный|Сетевой",   "",          0,0,0, "", ""),
    ("G_AREA",          "Площадь (м²)",                   "number",          "",                                         "150",       30, 2000, 10, "м²", ""),
    ("G_LICENSE",       "Лицензия",                       "radio",           "Есть|Планирую|Нет",                        "",          0,0,0, "", ""),
    ("G_KIDS_COUNT",    "Детей/учеников",                 "radio",           "8-15|16-30|31-60|60+",                     "",          0,0,0, "", ""),
    ("G_FEE",           "Ежемесячный платёж с одного (₸)","number",          "",                                         "70000",     5000, 500000, 1000, "₸", ""),
    ("G_STAFF_COUNT",   "Педагогов / инструкторов",       "radio",           "1-2|3-5|6-10|10+",                         "",          0,0,0, "чел", ""),
    ("G_FOOD_INCL",     "Питание включено в платёж",      "boolean",         "",                                         "",          0,0,0, "", "Для детсада"),
    ("G_FOOD_COST",     "Себест. питания на ребёнка/мес (₸)","number",       "",                                         "15000",     0, 100000, 500, "₸", ""),
    ("G_RENT",          "Аренда (₸/мес)",                 "number",          "",                                         "400000",    0, 5000000, 10000, "₸", ""),
    ("G_MARKETING",     "Маркетинг (₸/мес)",              "number",          "",                                         "80000",     0, 1000000, 5000, "₸", ""),
    ("G_SUMMER_DROP",   "Летний провал",                  "radio",           "0%|30%|60%|100%",                          "",          0,0,0, "%", ""),

    # ---------- ARCHETYPE H (mixed services) ----------
    ("H_FORMAT",        "Формат",                         "radio",           "Соло|Стандарт|Премиум|Сетевой",            "",          0,0,0, "", ""),
    ("H_LOC",           "Локация",                        "select",          "street|tc|residential_complex|own_building|business_center|home", "", 0,0,0, "", ""),
    ("H_AREA",          "Площадь (м²)",                   "number",          "",                                         "50",        0, 5000, 5, "м²", ""),
    ("H_CHECK",         "Средний чек / платёж (₸)",       "number",          "",                                         "5000",      100, 5000000, 100, "₸", ""),
    ("H_VOLUME",        "Клиентов / абонементов в месяц", "number",          "",                                         "100",       1, 10000, 10, "", ""),
    ("H_COGS",          "COGS (% от выручки)",            "radio",           "5%|15%|30%|50%|70%",                       "",          0,0,0, "%", "Аптека ~70, фитнес ~5"),
    ("H_RENT",          "Аренда (₸/мес)",                 "number",          "",                                         "250000",    0, 10000000, 10000, "₸", ""),
    ("H_MARKETING",     "Маркетинг (₸/мес)",              "number",          "",                                         "80000",     0, 2000000, 5000, "₸", ""),
    ("H_FOT_AVG",       "ФОТ сотрудника (₸/мес)",         "number",          "",                                         "200000",    0, 2000000, 10000, "₸", "Авто из 02_wages"),
    ("H_STAFF_COUNT",   "Команда",                        "radio",           "1|2-3|4-6|7-15|15+",                       "",          0,0,0, "чел", ""),

    # ---------- NICHE-SPECIFIC (B_*) ----------
    ("B_BARBER_TYPE",   "Формат барбершопа",              "radio",           "Мужской|Унисекс|Эконом-парикмахерская",    "",          0,0,0, "", ""),
    ("B_MANIC_SPEC",    "Специализация",                  "radio",           "Базовый|Наращивание|Педикюр|Всё",          "",          0,0,0, "", ""),
    ("B_MED_LIC",       "Мед.лицензия",                   "radio",           "Есть|Планирую|Нет",                        "",          0,0,0, "", ""),
    ("B_MASSAGE_TYPE",  "Тип массажа",                    "radio",           "Классический|Спортивный|Антицеллюлит|Лимфо","",         0,0,0, "", ""),
    ("B_BEAUTY_SVC",    "Услуги",                         "multi_select",    "Парикмахер|Маникюр|Косметолог|Массаж",     "",          0,0,0, "", ""),
    ("B_COFFEE_TYPE",   "Формат кофейни",                 "radio",           "Specialty|Mass-market|Островок",           "",          0,0,0, "", ""),
    ("B_PIZZA_FMT",     "Доставка",                       "radio",           "Своя|Агрегаторы|Только зал",               "",          0,0,0, "", ""),
    ("B_SUSHI_FMT",     "Кухня",                          "radio",           "Классическая|Комбо суши+пицца|Только доставка",          "", 0,0,0, "", ""),
    ("B_FASTFOOD_TYPE", "Концепция",                      "radio",           "Бургерная|Шаурма|Куриная|Стрит-фуд",       "",          0,0,0, "", ""),
    ("B_DONER_VAR",     "Меню",                           "radio",           "Только донер|+ кофе|+ лагман",             "",          0,0,0, "", ""),
    ("B_DRINKS_MIX",    "Напитки",                        "radio",           "Только бабл-ти|+ лимонады|+ кофе",         "",          0,0,0, "", ""),
    ("B_CANTEEN_TYPE",  "Локация столовой",               "radio",           "При заводе|БЦ|ТРЦ",                        "",          0,0,0, "", ""),
    ("B_BAKERY_TYPE",   "Тип пекарни",                    "radio",           "Свой|B2B-сеть|+ доставка",                 "",          0,0,0, "", ""),
    ("B_EVAL_TYPES",    "Виды оценки",                    "multi_select",    "Недвижимость|Транспорт|Бизнес",            "",          0,0,0, "", ""),
    ("B_ACC_PACKAGE",   "Тариф",                          "radio",           "Нулевой|1С|Полный аутсорс",                "",          0,0,0, "", ""),
    ("B_REAL_SEGMENT",  "Сегмент",                        "radio",           "Жильё|Коммерция|Аренда|Новостройки",       "",          0,0,0, "", ""),
    ("B_TIRE_HOTEL",    "Шинный отель",                   "boolean",         "",                                         "",          0,0,0, "", "Хранение шин"),
    ("B_WASH_TYPE",     "Тип мойки",                      "radio",           "Ручная|Самомойка|Робот",                   "",          0,0,0, "", ""),
    ("B_DETAIL_PKG",    "Пакеты",                         "radio",           "Только полировка|Комплекс|Премиум",        "",          0,0,0, "", ""),
    ("B_GROC_FMT",      "Формат точки",                   "radio",           "Киоск|Минимаркет|Супермаркет|Нишевый",     "",          0,0,0, "", ""),
    ("B_FV_SOURCE",     "Поставки",                       "radio",           "Рынок|Напрямую от фермера|Импорт",         "",          0,0,0, "", ""),
    ("B_MEAT_SLAUGH",   "Модель",                         "radio",           "Свой убой|Только переработка|Закуп туш",   "",          0,0,0, "", ""),
    ("B_FLOWER_SVC",    "Услуги",                         "multi_select",    "Букеты|Флористика|Доставка|Декор событий", "",          0,0,0, "", ""),
    ("B_PET_VET",       "Доп.услуги",                     "radio",           "Груминг|Ветаптека|Только корма",           "",          0,0,0, "", ""),
    ("B_OPT_DOC",       "Врач-офтальмолог",               "radio",           "Свой|По визиту|Нет",                       "",          0,0,0, "", ""),
    ("B_PARTS_SPEC",    "Специализация",                  "radio",           "Универсальный|Немецкие|Японские|Тюнинг",   "",          0,0,0, "", ""),
    ("B_BUILD_SVC",     "Доп.услуги",                     "radio",           "Доставка|Нет|+ услуги мастера",            "",          0,0,0, "", ""),
    ("B_FURN_TYPE",     "Тип мебели",                     "radio",           "Кухни|Шкафы-купе|Детская|Весь спектр",     "",          0,0,0, "", ""),
    ("B_LOFT_PROD",     "Продукция",                      "radio",           "Мебель|Металлоконструкции|Благоустройство","",          0,0,0, "", ""),
    ("B_WATER_PKG",     "Продукция",                      "radio",           "19л|Бутилированная|+ кулеры|+ доставка",   "",          0,0,0, "", ""),
    ("B_SEMI_TYPE",     "Ассортимент",                    "radio",           "Пельмени|Готовые блюда|Котлеты|Всё",       "",          0,0,0, "", ""),
    ("B_PRINT_TYPE",    "Тип типографии",                 "radio",           "Полиграфия|Широкий формат|Сувенирка|Весь спектр","",     0,0,0, "", ""),
    ("B_CONF_FMT",      "Формат кондитерки",              "radio",           "На заказ|Витрина|+ кафе|B2B-сети",         "",          0,0,0, "", ""),
    ("B_CATER_TYPE",    "Специализация кейтеринга",       "radio",           "Корпоративы|Банкеты|Детские|Выезд",        "",          0,0,0, "", ""),
    ("B_CARGO_FLEET",   "Машин в парке",                  "radio",           "1|2-3|5-10|10+",                           "",          0,0,0, "", ""),
    ("B_KG_AGE",        "Возрастная группа",              "radio",           "Ясли (1-3)|Средняя (3-5)|Весь спектр",     "",          0,0,0, "", ""),
    ("B_KIDS_FMT",      "Формат центра",                  "radio",           "Раннее развитие|Досадиковая группа|Подготовка к школе|Логопед|Lego/робототехника|Игровая в ТЦ","",0,0,0,"",""),
    ("B_LANG_FMT",      "Формат школы",                   "radio",           "Репетиторство|Мини-группы|Школа",          "",          0,0,0, "", ""),
    ("B_DRIVE_FLEET",   "Машин",                          "radio",           "2-3|4-6|7+",                               "",          0,0,0, "", ""),
    ("B_HOTEL_CLASS",   "Класс отеля",                    "radio",           "Хостел|2-3 звезды|4-5 звёзд|Апарт-отель",  "",          0,0,0, "", ""),
    ("B_HOTEL_ROOMS",   "Номеров",                        "radio",           "5-15|16-40|41-100|100+",                   "",          0,0,0, "", ""),
    ("B_HOTEL_OCC",     "Заполняемость",                  "radio",           "40%|60%|75%|85%+",                         "",          0,0,0, "%", ""),
    ("B_FIT_TYPE",      "Тип клуба",                      "radio",           "Тренажёрка|Групповые|Комплекс",            "",          0,0,0, "", ""),
    ("B_FIT_SUB_PRICE", "Средняя цена абонемента",        "number",          "",                                         "15000",     3000, 200000, 500, "₸", ""),
    ("B_YOGA_FMT",      "Формат",                         "radio",           "Студия|Выездной|Онлайн",                   "",          0,0,0, "", ""),
    ("B_CC_PCS",        "Кол-во ПК",                      "radio",           "10-20|21-40|41-80|80+",                    "",          0,0,0, "", ""),
    ("B_CC_RATE",       "Тариф (₸/час)",                  "number",          "",                                         "700",       100, 3000, 50, "₸", ""),
    ("B_PHARM_RX",      "Рецептурный отдел",              "boolean",         "",                                         "",          0,0,0, "", ""),
    ("B_DENT_INS",      "Работа со страховыми",           "boolean",         "",                                         "",          0,0,0, "", ""),
    ("B_FRANCHISE_PVZ", "Маркетплейсы",                   "multi_select",    "Wildberries|Kaspi Доставка|Ozon",          "",          0,0,0, "", ""),
    ("B_PVZ_EXTRA",     "Доп.услуги ПВЗ",                 "multi_select",    "Упаковка|Печать|Выдача документов",        "",          0,0,0, "", ""),
    ("B_REP_AVG_PRICE", "Средний чек ремонта (₸)",        "number",          "",                                         "6000",      500, 100000, 100, "₸", ""),
    ("B_REP_PARTS",     "Запчасти",                       "radio",           "Свои|Клиента",                             "",          0,0,0, "", ""),
    ("B_TAILOR_TYPE",   "Специализация",                  "radio",           "Мелкий ремонт|Пошив|Премиум",              "",          0,0,0, "", ""),
    ("B_PHOTO_FMT",     "Формат фотостудии",              "radio",           "Только съёмка|С интерьерами|Выездная|Школьная","",      0,0,0, "", ""),
    ("B_PHOTO_CHECK",   "Средний чек сессии (₸)",         "number",          "",                                         "25000",     3000, 300000, 500, "₸", ""),
    ("B_CLEAN_TYPE",    "Тип клининга",                   "radio",           "B2B офисы|B2C квартиры|Строительный",      "",          0,0,0, "", ""),
    ("B_DRY_TYPE",      "Тип химчистки",                  "radio",           "Приёмка|Полная химчистка",                 "",          0,0,0, "", ""),
    ("B_CARPET_PICKUP", "Выезд за ковром",                "boolean",         "",                                         "",          0,0,0, "", ""),
    ("B_LAUND_TYPE",    "Тип прачечной",                  "radio",           "Самообслуживание|Промышленная|B2B",        "",          0,0,0, "", ""),
]


# ---------------------------------------------------------------------------
# APPLICABILITY — which qids apply to which niche in which tier, and in what order
# Scheme: per-niche list is universal_block + archetype_block + niche_specific
# ---------------------------------------------------------------------------

# Universal block: order 1..10 (express + finmodel)
UNIVERSAL_EXPRESS = ["U_CITY", "U_NICHE", "U_CAPITAL"]
UNIVERSAL_FINMODEL = ["U_CITY", "U_NICHE", "U_CAPITAL",
                     "U_START_MONTH", "U_ENTITY",
                     "U_CREDIT", "U_CREDIT_AMOUNT", "U_CREDIT_RATE", "U_CREDIT_TERM",
                     "U_GROWTH_Y1"]

# Archetype → niches
ARCHETYPES = {
    "A": ["BARBER", "MANICURE", "BROW", "LASH", "SUGARING", "MASSAGE", "COSMETOLOGY", "BEAUTY"],
    "B": ["COFFEE", "PIZZA", "SUSHI", "FASTFOOD", "DONER", "BUBBLETEA", "CANTEEN", "BAKERY"],
    "C": ["NOTARY", "EVALUATION", "ACCOUNTING", "REALTOR"],
    "D": ["AUTOSERVICE", "TIRESERVICE", "CARWASH", "DETAILING"],
    "E": ["GROCERY", "FRUITSVEGS", "MEATSHOP", "FLOWERS", "PETSHOP", "OPTICS", "AUTOPARTS", "BUILDMAT"],
    "F": ["FURNITURE", "LOFTFURNITURE", "WATERPLANT", "SEMIFOOD", "PRINTING", "CONFECTION", "CATERING", "CARGO"],
    "G": ["KINDERGARTEN", "KIDSCENTER", "LANGUAGES", "DRIVING"],
    "H": ["CLEAN", "DRYCLEAN", "CARPETCLEAN", "LAUNDRY", "REPAIR_PHONE", "TAILOR", "PHOTO",
          "HOTEL", "FITNESS", "YOGA", "COMPCLUB", "PHARMACY", "DENTAL", "PVZ"],
}

# Archetype express blocks (in order after universal)
ARCHETYPE_EXPRESS = {
    "A": ["A_FORMAT", "A_LOC", "A_AREA", "A_CHAIRS", "A_STAFF_MODE", "A_CHECK"],
    "B": ["O_FORMAT", "O_LOC", "O_AREA", "O_CHECK", "O_TRAFFIC", "O_DELIVERY_SHARE", "O_STAFF_COUNT"],
    "C": ["P_FORMAT", "P_LICENSE", "P_STAFF", "P_CHECK_OR_RATE", "P_CLIENTS_PER_MONTH"],
    "D": ["D_FORMAT", "D_AREA", "D_POSTS", "D_CHECK", "D_LOAD", "D_STAFF_COUNT", "D_BRANDS"],
    "E": ["E_FORMAT", "E_LOC", "E_AREA", "E_TRAFFIC", "E_CHECK", "E_MARGIN", "E_STAFF_COUNT"],
    "F": ["F_FORMAT", "F_AREA", "F_CHECK_OR_UNIT", "F_VOLUME", "F_COGS", "F_STAFF_COUNT", "F_CLIENT_TYPE"],
    "G": ["G_FORMAT", "G_AREA", "G_LICENSE", "G_KIDS_COUNT", "G_FEE", "G_STAFF_COUNT"],
    "H": ["H_FORMAT", "H_LOC", "H_AREA", "H_CHECK", "H_VOLUME"],
}

# Archetype finmodel extensions (in order, appended AFTER express)
ARCHETYPE_FIN_EXTRA = {
    "A": ["A_LOAD", "A_COGS", "A_MARKETING", "A_FOT_MASTER", "A_RENT"],
    "B": ["O_FOODCOST", "O_AGGREGATOR_FEE", "O_RENT_MODE", "O_RENT_VAL", "O_MARKETING", "O_SEASON", "O_HOURS"],
    "C": ["P_RENT", "P_MARKETING", "P_FOT_SPEC", "P_REPEAT_RATE"],
    "D": ["D_COGS", "D_PARTS_MARKUP", "D_RENT", "D_MARKETING", "D_SEASON"],
    "E": ["E_INVENTORY_TURN", "E_LOSS_PCT", "E_INITIAL_STOCK", "E_RENT", "E_MARKETING", "E_HOURS"],
    "F": ["F_LEAD_TIME", "F_PREPAY_PCT", "F_EQUIPMENT_CAPEX", "F_RENT", "F_MARKETING", "F_LOGISTICS"],
    "G": ["G_FOOD_INCL", "G_FOOD_COST", "G_RENT", "G_MARKETING", "G_SUMMER_DROP"],
    "H": ["H_COGS", "H_RENT", "H_MARKETING", "H_FOT_AVG", "H_STAFF_COUNT"],
}

# Niche-specific questions (BOTH tiers)
NICHE_SPECIFIC = {
    "BARBER":       ["B_BARBER_TYPE"],
    "MANICURE":     ["B_MANIC_SPEC"],
    "COSMETOLOGY":  ["B_MED_LIC"],
    "MASSAGE":      ["B_MASSAGE_TYPE"],
    "BEAUTY":       ["B_BEAUTY_SVC"],
    "COFFEE":       ["B_COFFEE_TYPE"],
    "PIZZA":        ["B_PIZZA_FMT"],
    "SUSHI":        ["B_SUSHI_FMT"],
    "FASTFOOD":     ["B_FASTFOOD_TYPE"],
    "DONER":        ["B_DONER_VAR"],
    "BUBBLETEA":    ["B_DRINKS_MIX"],
    "CANTEEN":      ["B_CANTEEN_TYPE"],
    "BAKERY":       ["B_BAKERY_TYPE"],
    "EVALUATION":   ["B_EVAL_TYPES"],
    "ACCOUNTING":   ["B_ACC_PACKAGE"],
    "REALTOR":      ["B_REAL_SEGMENT"],
    "TIRESERVICE":  ["B_TIRE_HOTEL"],
    "CARWASH":      ["B_WASH_TYPE"],
    "DETAILING":    ["B_DETAIL_PKG"],
    "GROCERY":      ["B_GROC_FMT"],
    "FRUITSVEGS":   ["B_FV_SOURCE"],
    "MEATSHOP":     ["B_MEAT_SLAUGH"],
    "FLOWERS":      ["B_FLOWER_SVC"],
    "PETSHOP":      ["B_PET_VET"],
    "OPTICS":       ["B_OPT_DOC"],
    "AUTOPARTS":    ["B_PARTS_SPEC"],
    "BUILDMAT":     ["B_BUILD_SVC"],
    "FURNITURE":    ["B_FURN_TYPE"],
    "LOFTFURNITURE":["B_LOFT_PROD"],
    "WATERPLANT":   ["B_WATER_PKG"],
    "SEMIFOOD":     ["B_SEMI_TYPE"],
    "PRINTING":     ["B_PRINT_TYPE"],
    "CONFECTION":   ["B_CONF_FMT"],
    "CATERING":     ["B_CATER_TYPE"],
    "CARGO":        ["B_CARGO_FLEET"],
    "KINDERGARTEN": ["B_KG_AGE"],
    "KIDSCENTER":   ["B_KIDS_FMT"],
    "LANGUAGES":    ["B_LANG_FMT"],
    "DRIVING":      ["B_DRIVE_FLEET"],
    "HOTEL":        ["B_HOTEL_CLASS", "B_HOTEL_ROOMS", "B_HOTEL_OCC"],
    "FITNESS":      ["B_FIT_TYPE", "B_FIT_SUB_PRICE"],
    "YOGA":         ["B_YOGA_FMT"],
    "COMPCLUB":     ["B_CC_PCS", "B_CC_RATE"],
    "PHARMACY":     ["B_MED_LIC", "B_PHARM_RX"],
    "DENTAL":       ["B_MED_LIC", "B_DENT_INS"],
    "PVZ":          ["B_FRANCHISE_PVZ", "B_PVZ_EXTRA"],
    "REPAIR_PHONE": ["B_REP_AVG_PRICE", "B_REP_PARTS"],
    "TAILOR":       ["B_TAILOR_TYPE"],
    "PHOTO":        ["B_PHOTO_FMT", "B_PHOTO_CHECK"],
    "CLEAN":        ["B_CLEAN_TYPE"],
    "DRYCLEAN":     ["B_DRY_TYPE"],
    "CARPETCLEAN":  ["B_CARPET_PICKUP"],
    "LAUNDRY":      ["B_LAUND_TYPE"],
    # BROW, LASH, SUGARING, NOTARY — хватает архетипа без спец-вопросов
}


def _niche_to_arch(niche: str) -> str:
    for arch, niches in ARCHETYPES.items():
        if niche in niches:
            return arch
    return ""


def build_applicability_rows():
    out = []
    all_niches = set()
    for niches in ARCHETYPES.values():
        all_niches.update(niches)

    for niche_id in sorted(all_niches):
        arch = _niche_to_arch(niche_id)
        niche_specific = NICHE_SPECIFIC.get(niche_id, [])

        # Express
        express_qids = UNIVERSAL_EXPRESS + ARCHETYPE_EXPRESS.get(arch, []) + niche_specific
        for i, qid in enumerate(express_qids, start=1):
            out.append({
                "qid": qid, "niche_id": niche_id, "tier": "express",
                "order": i, "required": "yes",
            })

        # Finmodel — all express questions + finmodel extensions + rest of universal
        finmodel_qids = (
            UNIVERSAL_EXPRESS
            + ARCHETYPE_EXPRESS.get(arch, [])
            + ARCHETYPE_FIN_EXTRA.get(arch, [])
            + niche_specific
            + [q for q in UNIVERSAL_FINMODEL if q not in UNIVERSAL_EXPRESS]
        )
        # Deduplicate while preserving order
        seen = set(); ordered = []
        for q in finmodel_qids:
            if q not in seen:
                seen.add(q); ordered.append(q)
        for i, qid in enumerate(ordered, start=1):
            out.append({
                "qid": qid, "niche_id": niche_id, "tier": "finmodel",
                "order": i, "required": "yes",
            })
    return out


# ---------------------------------------------------------------------------
# DEPENDENCIES — conditional-show logic
# ---------------------------------------------------------------------------

DEPENDENCIES = [
    # Credit detail questions only show if U_CREDIT=yes
    ("U_CREDIT_AMOUNT", "U_CREDIT", "yes", "show"),
    ("U_CREDIT_RATE",   "U_CREDIT", "yes", "show"),
    ("U_CREDIT_TERM",   "U_CREDIT", "yes", "show"),
    # Aggregator fee only if delivery share > 0
    ("O_AGGREGATOR_FEE", "O_DELIVERY_SHARE", "!=0%", "show"),
    # Rent value depends on rent mode (only needed after mode picked)
    ("O_RENT_VAL",      "O_RENT_MODE", "any", "show"),
    # Food cost only if food is included in fee (kids)
    ("G_FOOD_COST",     "G_FOOD_INCL", "yes", "show"),
    # Parts markup only for AUTOSERVICE (не TIRE/WASH/DETAIL)
    ("D_PARTS_MARKUP",  "U_NICHE", "==AUTOSERVICE", "show"),
]


# ---------------------------------------------------------------------------
# WORKBOOK BUILDER
# ---------------------------------------------------------------------------

def write_workbook(out_path: str) -> None:
    wb = Workbook()

    header_fill = PatternFill(start_color="1F1F2E", end_color="1F1F2E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)

    # -------- Sheet 1: «Вопросы» --------
    ws = wb.active
    ws.title = "Вопросы"
    ws.cell(row=1, column=1, value="ZEREK — Каталог вопросов адаптивной анкеты v2").font = title_font
    ws.cell(row=2, column=1, value=f"Всего вопросов: {len(Q)}")
    ws.cell(row=3, column=1, value="input_type: select_remote | select_format | select | radio | number | range | boolean | multi_select")
    ws.cell(row=4, column=1, value="options: варианты через | (pipe)")

    headers = ["qid", "question_text", "input_type", "options", "placeholder",
               "min", "max", "step", "unit", "help"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=6, column=col_idx, value=h)
        c.font = header_font; c.fill = header_fill
    for i, row in enumerate(Q, start=7):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)

    widths = {1: 22, 2: 48, 3: 18, 4: 60, 5: 14, 6: 10, 7: 12, 8: 10, 9: 8, 10: 36}
    for col, w in widths.items():
        ws.column_dimensions[ws.cell(row=6, column=col).column_letter].width = w

    # -------- Sheet 2: «Применимость» --------
    ws2 = wb.create_sheet("Применимость")
    ws2.cell(row=1, column=1, value="ZEREK — Какие вопросы применяются к каким нишам в каком tier'е").font = title_font
    ws2.cell(row=2, column=1, value="tier: express (быстрая оценка) | finmodel (детальная)")
    ws2.cell(row=3, column=1, value="order: порядок рендера (возрастанию)")

    app_headers = ["qid", "niche_id", "tier", "order", "required"]
    for col_idx, h in enumerate(app_headers, start=1):
        c = ws2.cell(row=6, column=col_idx, value=h)
        c.font = header_font; c.fill = header_fill
    rows = build_applicability_rows()
    for i, row in enumerate(rows, start=7):
        for j, key in enumerate(app_headers, start=1):
            ws2.cell(row=i, column=j, value=row[key])
    widths2 = {1: 24, 2: 16, 3: 12, 4: 8, 5: 12}
    for col, w in widths2.items():
        ws2.column_dimensions[ws2.cell(row=6, column=col).column_letter].width = w

    # -------- Sheet 3: «Зависимости» --------
    ws3 = wb.create_sheet("Зависимости")
    ws3.cell(row=1, column=1, value="ZEREK — Условный показ вопросов").font = title_font
    ws3.cell(row=2, column=1, value="Если condition не выполнено — вопрос скрыт и пропущен из анкеты.")
    ws3.cell(row=3, column=1, value="condition: '=={val}' | '!={val}' | 'yes' | 'any' (любое значение).")
    dep_headers = ["qid", "depends_on", "condition", "action"]
    for col_idx, h in enumerate(dep_headers, start=1):
        c = ws3.cell(row=6, column=col_idx, value=h)
        c.font = header_font; c.fill = header_fill
    for i, (qid, dep, cond, act) in enumerate(DEPENDENCIES, start=7):
        ws3.cell(row=i, column=1, value=qid)
        ws3.cell(row=i, column=2, value=dep)
        ws3.cell(row=i, column=3, value=cond)
        ws3.cell(row=i, column=4, value=act)
    widths3 = {1: 22, 2: 22, 3: 28, 4: 10}
    for col, w in widths3.items():
        ws3.column_dimensions[ws3.cell(row=6, column=col).column_letter].width = w

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"✅ {out_path}")
    print(f"   Вопросов: {len(Q)}")
    print(f"   Применимость: {len(rows)} строк (58 ниш × 2 tier)")
    print(f"   Зависимостей: {len(DEPENDENCIES)}")


if __name__ == "__main__":
    repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out = os.path.join(repo_root, "data", "kz", "09_surveys.xlsx")
    write_workbook(out)
